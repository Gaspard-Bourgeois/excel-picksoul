"""
analyseHypertrempe.py

Analyse un (ou plusieurs) fichier(s) Excel d'enregistreur de température
(type SMARTDAC+, colonnes CHxxxx / CHCxxx) et génère, pour chacun, une
image regroupant :
  - un graphique gauche (pleine hauteur) montrant l'ensemble des voies
    retenues après filtrage (voies grisées, courbes "première"/"dernière"
    surlignées),
  - 4 graphiques à droite, un par phase du cycle (palier 1, palier 2,
    refroidissement palier 3, sortie finale palier 4), chacun annoté des
    points d'intérêt qui le concernent,
ainsi qu'un second fichier texte contenant les valeurs de synthèse du
cycle (température initiale, vitesses de montée/descente, durées de
regroupement/maintien/sortie, température de maintien 2), au format
imposé par un exemple fourni (voir hypothèse 8).

Les paramètres se trouvent dans un fichier YAML (par défaut
analyseHypertrempe.yaml), avec quelques surcharges possibles en ligne
de commande pour les seuils principaux.

Usage :
    python analyseHypertrempe.py fichier1.xlsx fichier2.xlsx
    (sans fichier -> traite tous les .xlsx du dossier courant, hors fichiers
    déjà suffixés)
    (paramètres absents de la ligne de commande -> lus dans
    analyseHypertrempe.yaml)

-------------------------------------------------------------------------
HYPOTHÈSES / CHOIX DE CONCEPTION (assumés faute de spécification exacte) :

1. Structure du fichier Excel : la ligne d'en-tête des voies est repérée
   automatiquement en cherchant la première ligne contenant des cellules du
   type "CH0001", "CH0002"... ou "CHC001", "CHC002"... (pas de numéro de
   ligne fixe). La ligne de début des données est repérée en cherchant,
   après cette ligne, la ligne dont la colonne A vaut "Date" et la colonne
   B vaut "Time" : les données commencent juste après. La lecture s'arrête
   à la première ligne dont la colonne "Date" est vide. Aucune correction
   de chronologie n'est appliquée : les dates/heures du fichier sont
   utilisées telles quelles.

2. Cellule `titre_graphique` : repérée en cherchant, en colonne A, la ligne
   dont le libellé vaut exactement "Batch No." (fixe), et en lisant la
   valeur en colonne C (fixe également — ces deux réglages ne sont plus
   exposés dans le YAML, à la demande explicite d'une itération
   précédente).

3. `CH_ignore_dT` : une voie est CONSERVÉE si (max - min) de ses valeurs
   sur l'ensemble du fichier est STRICTEMENT SUPÉRIEUR à `CH_ignore_dT` ;
   sinon elle est ignorée (voie trop stable / non pertinente). Une voie
   contenant au moins une valeur de type "-OVER" est ignorée avant même ce
   calcul, quelle que soit son amplitude.

4. "Première courbe" / "dernière courbe" : on utilise TOUTES les voies non
   filtrées pour le tracé (voir hypothèse précédente sur `CH_ignore_dT`),
   mais la détection des instants s'appuie sur DEUX voies fixes, désignées
   par leur position dans l'ordre des colonnes du fichier Excel :
     - la "première courbe" = la première voie retenue (première colonne
       CHxxxx non filtrée),
     - la "dernière courbe" = la dernière voie retenue (dernière colonne
       CHxxxx non filtrée).
   C'est l'interprétation retenue pour "la première/dernière courbe" du
   cahier des charges (les instants de "regroupement" mesurent alors
   l'écart entre le point du chargement qui chauffe/refroidit le plus vite
   — la première voie — et celui qui chauffe/refroidit le plus lentement —
   la dernière voie). Si une seule voie est retenue après filtrage, les
   deux coïncident (durées de regroupement nulles).

5. Détection des instants (recherche séquentielle sur les DEUX courbes
   première/dernière ; chaque étape reprend la recherche à partir de
   l'instant trouvé à l'étape précédente) :
     - t0 (début montée) = 1ère courbe franchit (montant)
       température initiale + `temp_delta_ambiant`
     - t1 (fin montée 1) = 1ère courbe franchit (montant)
       `temp_palier_1` - `temp_delta_1` -> vitesse de montée 1 (°C/h)
       calculée entre t0 et t1
     - t2 (regroupement 1) = dernière courbe franchit (montant) le même
       seuil que t1 -> durée de regroupement 1 = t2 - t1
     - t3 (maintien 1) = 1ère courbe franchit (montant)
       `temp_palier_1` + `temp_delta_1` -> durée de maintien 1 = t3 - t2
     - t4 (fin montée 2) = 1ère courbe franchit (montant)
       `temp_palier_2` - `temp_delta_2` -> vitesse de montée 2 (°C/h)
       calculée entre t3 et t4
     - t5 (regroupement 2) = dernière courbe franchit (montant) le même
       seuil que t4 -> durée de regroupement 2 = t5 - t4
       [ÉTAPE AJOUTÉE : le cahier des charges ne décrivait pas
       explicitement de palier "regroupement 2" symétrique du
       regroupement 1, mais la liste des 12 valeurs de sortie demandées
       (hypothèse 8) inclut une "Durée de regroupement 2" — cette étape a
       donc été ajoutée par symétrie avec le palier 1, entre la fin de
       montée 2 et le début du maintien 2]
     - t6 (maintien 2) = 1ère courbe franchit (montant)
       `temp_palier_2` + `temp_delta_2` -> durée de maintien 2 = t6 - t5 ;
       la "température de maintien 2" est le maximum atteint, TOUTES
       voies retenues confondues, entre t5 et t6
     - t7 (sortie 2) = dernière courbe franchit (descendant)
       `temp_palier_2` - `temp_delta_2` -> durée de sortie 2 = t7 - t6
     - t8 (fin descente 3) = 1ère courbe franchit (descendant)
       `temp_palier_3` + `temp_delta_3` -> vitesse de descente 3 (°C/h)
       calculée entre t7 et t8
     - t9 (regroupement 3) = dernière courbe franchit (descendant)
       `temp_palier_3` - `temp_delta_3` -> durée de regroupement 3 = t9 - t8
     - t10 (fin descente 4) = 1ère courbe franchit (descendant)
       `temp_palier_4` -> vitesse de descente 4 (°C/h) calculée entre t9
       et t10
   Tous les franchissements sont interpolés linéairement entre les deux
   échantillons encadrants. Si une étape n'est pas trouvée, les étapes
   suivantes ne sont pas calculées (valeurs "N/A" dans les sorties).

6. Fenêtres de zoom : deux paramètres uniques `zoom_marge_temperature`
   (°C) et `zoom_marge_temps` (minutes), appliqués aux 4 graphiques de
   droite, ajoutés de part et d'autre de la plage utile de chaque phase.

7. Fichier de sortie image : `<nom_fichier_excel>_<suffixe>.png`.
   Fichier de sortie résultats : `<nom_fichier_excel>_<suffixe_resultats>.txt`.
   (les extensions ne sont plus paramétrables, voir hypothèse 2).

8. Formalisme STRICT du fichier de résultats (imposé, reproduit tel quel,
   y compris la ponctuation/orthographe des libellés fournis en exemple) :
   12 lignes de libellés puis 12 lignes de valeurs, dans cet ordre :
       température initiale                    (tabulation AVANT la valeur)
       vitesse de montée 1                      (avant)
       durée de regroupement 1                  (avant)
       durée du maintient 1                     (avant)
       Vitesse de montée 2                      (APRÈS la valeur)
       Durée de regroupement 2                  (après)
       Température de maintient 2               (après)
       Durée du maintient 2                     (après)
       Durée de la sortie 2                     (après)
       Vitesse de descente 3                    (après)
       Durée du regroupement 3                  (avant)
       vitese de descente 4                     (avant)
   soit, pour les valeurs (exemple donné) :
       \t12°C
       \t20°C/h
       \t5min
       \t40min
       20°C/h\t
       5min\t
       40°C\t
       3h02min\t
       2min\t
       20°C/h\t
       \t2min
       \t10°C/h
   Le fichier final concatène le bloc des 12 libellés puis le bloc des 12
   valeurs, chacun respectant la position de tabulation indiquée.

9. Température initiale : première valeur (au premier instant du fichier)
   de la "première courbe" (voir hypothèse 4).

10. Toutes les valeurs de sortie sont arrondies à l'entier le plus proche,
    sans décimale. Durées formatées sans espace, en minutes en dessous de
    60 min ("5min"), sinon en heures+minutes ("3h02min"). Les vitesses de
    montée/descente sont exprimées en °C/h (et non plus en °C/min).

11. Couleurs : chaque point d'intérêt (t0 à t10) a sa propre couleur fixe,
    réutilisée à l'identique sur tous les graphiques (trait de seuil,
    trait vertical, marqueur, texte d'annotation). Les annotations sont
    posées sur fond blanc semi-opaque, et leur position (au-dessus/en
    dessous, à gauche/à droite du point) est choisie automatiquement selon
    la position du point dans la fenêtre affichée, pour rester à
    l'intérieur du graphique.
-------------------------------------------------------------------------
"""

import argparse
import os
import re
import sys
from datetime import datetime, timedelta

import matplotlib
matplotlib.use("Agg")
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import openpyxl
import yaml


# =========================================================================
# Constantes non paramétrables (retirées du YAML à la demande de l'utilisateur)
# =========================================================================

FORMAT_IMAGE = 'png'
FORMAT_RESULTATS = 'txt'
FEUILLE = None
TITRE_GRAPHIQUE_COL = 3
TITRE_GRAPHIQUE_LABEL = "Batch No."

DEFAULT_CONFIG_FILE = "analyseHypertrempe.yaml"


# =========================================================================
# Lecture du fichier Excel
# =========================================================================

_RE_CHANNEL = re.compile(r'^CH[A-Z]?\d{3,4}$', re.IGNORECASE)


def trouver_ligne_entete_channels(ws, max_rows_scan=200):
    """Cherche la première ligne contenant des libellés de voie CHxxxx /
    CHCxxx. Renvoie (numero_ligne, {colonne: nom_voie})."""
    for r in range(1, min(max_rows_scan, ws.max_row) + 1):
        canaux = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and _RE_CHANNEL.match(v.strip()):
                canaux[c] = v.strip()
        if canaux:
            return r, canaux
    raise ValueError("Impossible de trouver la ligne d'en-tête des voies (CHxxxx / CHCxxx).")


def trouver_ligne_debut_donnees(ws, ligne_apres, max_rows_scan=50):
    """Cherche, après la ligne d'en-tête des voies, la ligne 'Date'/'Time' et
    renvoie le numéro de la première ligne de données (juste après)."""
    borne = min(ligne_apres + max_rows_scan, ws.max_row)
    for r in range(ligne_apres, borne + 1):
        v1 = ws.cell(row=r, column=1).value
        v2 = ws.cell(row=r, column=2).value
        if isinstance(v1, str) and v1.strip().lower() == 'date' and \
                isinstance(v2, str) and v2.strip().lower() == 'time':
            return r + 1
    raise ValueError("Impossible de trouver la ligne d'en-tête 'Date' / 'Time'.")


def trouver_titre_graphique(ws, max_rows_scan=200):
    for r in range(1, min(max_rows_scan, ws.max_row) + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == TITRE_GRAPHIQUE_LABEL:
            val = ws.cell(row=r, column=TITRE_GRAPHIQUE_COL).value
            return "" if val is None else str(val).strip()
    return ""


def parser_date_heure(val_date, val_time):
    if isinstance(val_date, datetime):
        d = val_date.date()
    else:
        d = datetime.strptime(str(val_date).strip(), "%Y/%m/%d").date()

    if hasattr(val_time, 'hour'):
        t = val_time.time() if isinstance(val_time, datetime) else val_time
    else:
        s = str(val_time).strip()
        fmt = "%H:%M:%S" if s.count(':') == 2 else "%H:%M"
        t = datetime.strptime(s, fmt).time()

    return datetime.combine(d, t)


def lire_donnees_excel(fichier):
    wb = openpyxl.load_workbook(fichier, data_only=True)
    ws = wb[FEUILLE] if FEUILLE else wb[wb.sheetnames[0]]

    ligne_ch, canaux_col = trouver_ligne_entete_channels(ws)
    ligne_debut = trouver_ligne_debut_donnees(ws, ligne_ch)
    titre = trouver_titre_graphique(ws)

    dates = []
    valeurs_brutes = {nom: [] for nom in canaux_col.values()}

    r = ligne_debut
    while True:
        v_date = ws.cell(row=r, column=1).value
        if v_date is None or str(v_date).strip() == '':
            break
        v_time = ws.cell(row=r, column=2).value
        dates.append(parser_date_heure(v_date, v_time))
        for c, nom in canaux_col.items():
            valeurs_brutes[nom].append(ws.cell(row=r, column=c).value)
        r += 1

    if not dates:
        raise ValueError(f"Aucune donnée trouvée dans {fichier}.")

    return titre, dates, valeurs_brutes


# =========================================================================
# Filtrage des voies
# =========================================================================

def _est_valeur_over(v):
    return isinstance(v, str) and 'OVER' in v.upper()


def filtrer_voies(valeurs_brutes, ch_ignore_dT):
    """Retire les voies contenant une valeur de type -OVER, puis celles dont
    l'amplitude (max-min) est <= ch_ignore_dT. Renvoie {nom: [floats]},
    en conservant l'ordre des colonnes du fichier Excel d'origine."""
    voies = {}
    for nom, valeurs in valeurs_brutes.items():
        if any(_est_valeur_over(v) for v in valeurs):
            continue
        try:
            floats = [float(v) for v in valeurs]
        except (TypeError, ValueError):
            continue
        if (max(floats) - min(floats)) > ch_ignore_dT:
            voies[nom] = floats
    return voies


# =========================================================================
# Détection des instants de franchissement de seuil
# =========================================================================

def detecter_franchissement(dates, valeurs, seuil, sens, idx_debut=0):
    """sens = 'montant' (franchissement vers le haut) ou 'descente'
    (franchissement vers le bas). Recherche à partir de l'indice idx_debut.
    Renvoie (instant interpolé, indice de l'échantillon suivant le
    franchissement) ou (None, None) si le seuil n'est jamais franchi."""
    depart = max(1, idx_debut)
    for i in range(depart, len(valeurs)):
        v0, v1 = valeurs[i - 1], valeurs[i]
        if sens == 'montant':
            franchi = v0 < seuil <= v1
        else:
            franchi = v0 >= seuil > v1
        if not franchi:
            continue
        t0, t1 = dates[i - 1], dates[i]
        frac = 0.0 if v1 == v0 else (seuil - v0) / (v1 - v0)
        frac = max(0.0, min(1.0, frac))
        return t0 + (t1 - t0) * frac, i
    return None, None


def calculer_cycle(dates, voies, config):
    """Calcule les instants t0..t10 et les valeurs de synthèse du cycle
    d'hypertrempe (voir hypothèse 5 en tête de fichier). Renvoie un dict."""
    noms = list(voies)
    premiere_nom = noms[0]
    derniere_nom = noms[-1]
    premiere = voies[premiere_nom]
    derniere = voies[derniere_nom]

    cycle = {f't{i}': None for i in range(11)}
    cycle.update({f'seuil_t{i}': None for i in range(11)})
    cycle.update({
        'premiere_nom': premiere_nom,
        'derniere_nom': derniere_nom,
        'temp_initiale': premiere[0] if premiere else None,
        'vitesse_montee_1': None,
        'duree_regroupement_1_min': None,
        'duree_maintien_1_min': None,
        'vitesse_montee_2': None,
        'duree_regroupement_2_min': None,
        'temp_maintien_2': None,
        'instant_maintien_2': None,
        'duree_maintien_2_min': None,
        'duree_sortie_2_min': None,
        'vitesse_descente_3': None,
        'duree_regroupement_3_min': None,
        'vitesse_descente_4': None,
    })

    if cycle['temp_initiale'] is None:
        return cycle

    seuil_t0 = cycle['temp_initiale'] + config['temp_delta_ambiant']
    cycle['seuil_t0'] = seuil_t0
    t0, i0 = detecter_franchissement(dates, premiere, seuil_t0, 'montant', 0)
    cycle['t0'] = t0
    if t0 is None:
        return cycle

    seuil_t1 = config['temp_palier_1'] - config['temp_delta_1']
    cycle['seuil_t1'] = seuil_t1
    t1, i1 = detecter_franchissement(dates, premiere, seuil_t1, 'montant', i0)
    cycle['t1'] = t1
    if t1 is None:
        return cycle
    heures = (t1 - t0).total_seconds() / 3600.0
    if heures > 0:
        cycle['vitesse_montee_1'] = (seuil_t1 - seuil_t0) / heures

    cycle['seuil_t2'] = seuil_t1
    t2, i2 = detecter_franchissement(dates, derniere, seuil_t1, 'montant', i1)
    cycle['t2'] = t2
    if t2 is None:
        return cycle
    cycle['duree_regroupement_1_min'] = (t2 - t1).total_seconds() / 60.0

    seuil_t3 = config['temp_palier_1'] + config['temp_delta_1']
    cycle['seuil_t3'] = seuil_t3
    t3, i3 = detecter_franchissement(dates, premiere, seuil_t3, 'montant', i2)
    cycle['t3'] = t3
    if t3 is None:
        return cycle
    cycle['duree_maintien_1_min'] = (t3 - t2).total_seconds() / 60.0

    seuil_t4 = config['temp_palier_2'] - config['temp_delta_2']
    cycle['seuil_t4'] = seuil_t4
    t4, i4 = detecter_franchissement(dates, premiere, seuil_t4, 'montant', i3)
    cycle['t4'] = t4
    if t4 is None:
        return cycle
    heures = (t4 - t3).total_seconds() / 3600.0
    if heures > 0:
        cycle['vitesse_montee_2'] = (seuil_t4 - seuil_t3) / heures

    cycle['seuil_t5'] = seuil_t4
    t5, i5 = detecter_franchissement(dates, derniere, seuil_t4, 'montant', i4)
    cycle['t5'] = t5
    if t5 is None:
        return cycle
    cycle['duree_regroupement_2_min'] = (t5 - t4).total_seconds() / 60.0

    seuil_t6 = config['temp_palier_2'] + config['temp_delta_2']
    cycle['seuil_t6'] = seuil_t6
    t6, i6 = detecter_franchissement(dates, premiere, seuil_t6, 'montant', i5)
    cycle['t6'] = t6
    if t6 is None:
        return cycle
    cycle['duree_maintien_2_min'] = (t6 - t5).total_seconds() / 60.0

    debut_idx, fin_idx = min(i5, i6), max(i5, i6)
    segment_dates = dates[debut_idx:fin_idx + 1]
    meilleur_valeur, meilleur_instant = None, None
    for valeurs in voies.values():
        segment = valeurs[debut_idx:fin_idx + 1]
        for t, v in zip(segment_dates, segment):
            if meilleur_valeur is None or v > meilleur_valeur:
                meilleur_valeur, meilleur_instant = v, t
    cycle['temp_maintien_2'] = meilleur_valeur
    cycle['instant_maintien_2'] = meilleur_instant

    seuil_t7 = config['temp_palier_2'] - config['temp_delta_2']
    cycle['seuil_t7'] = seuil_t7
    t7, i7 = detecter_franchissement(dates, derniere, seuil_t7, 'descente', i6)
    cycle['t7'] = t7
    if t7 is None:
        return cycle
    cycle['duree_sortie_2_min'] = (t7 - t6).total_seconds() / 60.0

    seuil_t8 = config['temp_palier_3'] + config['temp_delta_3']
    cycle['seuil_t8'] = seuil_t8
    t8, i8 = detecter_franchissement(dates, premiere, seuil_t8, 'descente', i7)
    cycle['t8'] = t8
    if t8 is None:
        return cycle
    heures = (t8 - t7).total_seconds() / 3600.0
    if heures > 0:
        cycle['vitesse_descente_3'] = (seuil_t7 - seuil_t8) / heures

    seuil_t9 = config['temp_palier_3'] - config['temp_delta_3']
    cycle['seuil_t9'] = seuil_t9
    t9, i9 = detecter_franchissement(dates, derniere, seuil_t9, 'descente', i8)
    cycle['t9'] = t9
    if t9 is None:
        return cycle
    cycle['duree_regroupement_3_min'] = (t9 - t8).total_seconds() / 60.0

    seuil_t10 = config['temp_palier_4']
    cycle['seuil_t10'] = seuil_t10
    t10, i10 = detecter_franchissement(dates, premiere, seuil_t10, 'descente', i9)
    cycle['t10'] = t10
    if t10 is None:
        return cycle
    heures = (t10 - t9).total_seconds() / 3600.0
    if heures > 0:
        cycle['vitesse_descente_4'] = (seuil_t9 - seuil_t10) / heures

    return cycle


# =========================================================================
# Formatage
# =========================================================================

def fmt_num(x, decimales=0):
    if x is None:
        return "N/A"
    s = f"{x:.{decimales}f}"
    if decimales > 0 and s.endswith('.' + '0' * decimales):
        s = s.split('.')[0]
    return s


def fmt_temperature(x):
    return "N/A" if x is None else f"{fmt_num(x)}°C"


def fmt_duree(minutes):
    """Durée sans espace : "5min" en dessous de 60 min, sinon "3h02min"."""
    if minutes is None:
        return "N/A"
    total = round(minutes)
    signe = '-' if total < 0 else ''
    total_abs = abs(total)
    if total_abs >= 60:
        h, m = divmod(total_abs, 60)
        return f"{signe}{h}h{m:02d}min"
    return f"{signe}{total_abs}min"


def fmt_vitesse_h(x):
    return "N/A" if x is None else f"{fmt_num(x)}°C/h"


# =========================================================================
# Graphique
# =========================================================================

_COULEUR_GRISE = '0.75'
_COULEUR_PREMIERE = 'black'
_COULEUR_DERNIERE = '#8c1616'
_COULEUR_INITIALE = '#8c564b'
_COULEUR_MAX_MAINTIEN = 'teal'

_ORDRE_POINTS = [f't{i}' for i in range(11)]
_PALETTE = plt.get_cmap('tab20').colors
COULEURS_POINTS = {cle: _PALETTE[i] for i, cle in enumerate(_ORDRE_POINTS)}
LABELS_POINTS = {
    't0': 'Début montée',
    't1': 'Fin montée 1',
    't2': 'Regroupement 1',
    't3': 'Maintien 1',
    't4': 'Fin montée 2',
    't5': 'Regroupement 2',
    't6': 'Maintien 2',
    't7': 'Sortie 2',
    't8': 'Fin descente 3',
    't9': 'Regroupement 3',
    't10': 'Fin descente 4',
}


def _tracer_courbes_fenetre(ax, dates, voies, premiere_nom, derniere_nom, t_min, t_max):
    """Trace, dans la fenêtre [t_min, t_max], toutes les voies retenues en
    gris (une seule entrée de légende "Autres voies"), et les courbes
    première/dernière surlignées."""
    premiere_grise = True
    for nom, valeurs in voies.items():
        if nom in (premiere_nom, derniere_nom):
            continue
        xs, ys = [], []
        for t, v in zip(dates, valeurs):
            if t_min <= t <= t_max:
                xs.append(t)
                ys.append(v)
        if xs:
            label = "Autres voies" if premiere_grise else None
            ax.plot(xs, ys, color=_COULEUR_GRISE, linewidth=0.8, alpha=0.9, zorder=1, label=label)
            premiere_grise = False

    voies_mises_en_avant = [(premiere_nom, _COULEUR_PREMIERE, f"{premiere_nom} (première)")]
    if derniere_nom != premiere_nom:
        voies_mises_en_avant.append((derniere_nom, _COULEUR_DERNIERE, f"{derniere_nom} (dernière)"))
    for nom, couleur, label in voies_mises_en_avant:
        xs = [t for t in dates if t_min <= t <= t_max]
        ys = [v for t, v in zip(dates, voies[nom]) if t_min <= t <= t_max]
        if xs:
            ax.plot(xs, ys, color=couleur, linewidth=2, label=label, zorder=3)


def _annoter_point(ax, instant, valeur, texte, couleur, x_range, y_range, label=None):
    """Place un marqueur + une étiquette pour un point d'intérêt. La
    position de l'étiquette (haut/bas, gauche/droite) est choisie
    automatiquement selon la position du point dans (x_range, y_range),
    pour rester à l'intérieur du graphique. x_range peut contenir des
    datetime (fenêtres calculées manuellement) ou des floats matplotlib
    (ex. ax.get_xlim()) : les deux sont normalisés avant comparaison."""
    def _num(x):
        return mdates.date2num(x) if isinstance(x, datetime) else x

    x_min, x_max = x_range
    y_min, y_max = y_range
    a_gauche = _num(instant) <= (_num(x_min) + _num(x_max)) / 2
    au_dessus = valeur <= (y_min + y_max) / 2
    dx = 8 if a_gauche else -8
    dy = 8 if au_dessus else -8
    ha = 'left' if a_gauche else 'right'
    va = 'bottom' if au_dessus else 'top'
    ax.plot([instant], [valeur], marker='o', color=couleur, markersize=6,
             markeredgecolor='white', markeredgewidth=0.7, zorder=6, label=label,
             clip_on=True)
    ax.annotate(texte, xy=(instant, valeur), xytext=(dx, dy), textcoords='offset points',
                 color=couleur, fontsize=8, fontweight='bold', ha=ha, va=va, clip_on=True,
                 bbox=dict(boxstyle='round,pad=0.2', facecolor='white',
                           edgecolor=couleur, linewidth=1, alpha=0.9))


def _tracer_zoom_phase(ax, dates, voies, premiere_nom, derniere_nom, config, points, titre):
    """points : liste de tuples (instant, valeur, couleur, label, ligne_horizontale)."""
    if not points or any(p[0] is None or p[1] is None for p in points):
        ax.set_title(f"{titre} : seuils non atteints", fontweight='bold', pad=8)
        ax.text(0.5, 0.5, "Seuils non franchis", ha='center', va='center',
                 transform=ax.transAxes, color='dimgray')
        return

    marge_t = timedelta(minutes=config['zoom_marge_temps'])
    marge_T = config['zoom_marge_temperature']
    instants = [p[0] for p in points]
    valeurs = [p[1] for p in points]
    t_min, t_max = min(instants) - marge_t, max(instants) + marge_t
    y_min, y_max = min(valeurs) - marge_T, max(valeurs) + marge_T
    x_range, y_range = (t_min, t_max), (y_min, y_max)

    _tracer_courbes_fenetre(ax, dates, voies, premiere_nom, derniere_nom, t_min, t_max)

    for instant, valeur, couleur, label, ligne_horizontale in points:
        if ligne_horizontale:
            ax.axhline(valeur, color=couleur, linestyle='--', linewidth=1.2)
        _annoter_point(ax, instant, valeur,
                        f"{instant.strftime('%H:%M:%S')}\n{fmt_temperature(valeur)}",
                        couleur, x_range, y_range, label=label)

    ax.set_xlim(t_min, t_max)
    ax.set_ylim(y_min, y_max)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
    for lbl in ax.get_xticklabels():
        lbl.set_rotation(20)
        lbl.set_ha('right')
    ax.set_ylabel("Température (°C)")
    ax.grid(True, alpha=0.3)
    ax.set_title(titre, fontweight='bold', pad=8)


def tracer_graphique(fichier, titre, dates, voies, cycle, config):
    premiere_nom, derniere_nom = cycle['premiere_nom'], cycle['derniere_nom']

    fig = plt.figure(figsize=tuple(config['figure_taille']))
    gs = fig.add_gridspec(4, 2, width_ratios=[1.4, 1], hspace=0.7, wspace=0.25)
    ax_gauche = fig.add_subplot(gs[:, 0])
    ax_p1 = fig.add_subplot(gs[0, 1])
    ax_p2 = fig.add_subplot(gs[1, 1])
    ax_p3 = fig.add_subplot(gs[2, 1])
    ax_p4 = fig.add_subplot(gs[3, 1])

    # --- Graphique global (gauche, pleine hauteur) ---
    premiere_grise = True
    for nom, valeurs in voies.items():
        if nom in (premiere_nom, derniere_nom):
            continue
        label = "Autres voies" if premiere_grise else None
        ax_gauche.plot(dates, valeurs, color=_COULEUR_GRISE, linewidth=0.8,
                        alpha=0.9, label=label, zorder=1)
        premiere_grise = False
    ax_gauche.plot(dates, voies[premiere_nom], color=_COULEUR_PREMIERE, linewidth=2,
                    label=f"{premiere_nom} (première)", zorder=3)
    if derniere_nom != premiere_nom:
        ax_gauche.plot(dates, voies[derniere_nom], color=_COULEUR_DERNIERE, linewidth=2,
                        label=f"{derniere_nom} (dernière)", zorder=3)

    if cycle['temp_initiale'] is not None and dates:
        x_range_gauche = ax_gauche.get_xlim()
        y_range_gauche = ax_gauche.get_ylim()
        _annoter_point(ax_gauche, dates[0], cycle['temp_initiale'],
                        f"{dates[0].strftime('%H:%M:%S')}\n{fmt_temperature(cycle['temp_initiale'])}",
                        _COULEUR_INITIALE, x_range_gauche, y_range_gauche,
                        label="Température initiale")

    for cle in _ORDRE_POINTS:
        instant = cycle[cle]
        if instant is not None:
            ax_gauche.axvline(instant, color=COULEURS_POINTS[cle], linestyle='--',
                                linewidth=1.2, alpha=0.8, label=LABELS_POINTS[cle])

    ax_gauche.set_xlabel("Temps")
    ax_gauche.set_ylabel("Température (°C)")
    ax_gauche.xaxis.set_major_formatter(mdates.DateFormatter('%Y/%m/%d %H:%M:%S'))
    for lbl in ax_gauche.get_xticklabels():
        lbl.set_rotation(30)
        lbl.set_ha('right')
    ax_gauche.grid(True, alpha=0.3)
    ax_gauche.legend(loc='upper left', ncol=3, fontsize=6, framealpha=0.9)

    # --- Phase 1 : palier 1 (montée / regroupement / maintien) ---
    titre_p1 = (f"Palier 1 — montée {fmt_vitesse_h(cycle['vitesse_montee_1'])} | "
                f"regroup. {fmt_duree(cycle['duree_regroupement_1_min'])} | "
                f"maintien {fmt_duree(cycle['duree_maintien_1_min'])}")
    points_p1 = [
        (cycle['t0'], cycle['seuil_t0'], COULEURS_POINTS['t0'], LABELS_POINTS['t0'], True),
        (cycle['t1'], cycle['seuil_t1'], COULEURS_POINTS['t1'], LABELS_POINTS['t1'], True),
        (cycle['t2'], cycle['seuil_t2'], COULEURS_POINTS['t2'], LABELS_POINTS['t2'], True),
        (cycle['t3'], cycle['seuil_t3'], COULEURS_POINTS['t3'], LABELS_POINTS['t3'], True),
    ]
    _tracer_zoom_phase(ax_p1, dates, voies, premiere_nom, derniere_nom, config, points_p1, titre_p1)

    # --- Phase 2 : palier 2 (montée / regroupement / maintien / sortie) ---
    titre_p2 = (f"Palier 2 — montée {fmt_vitesse_h(cycle['vitesse_montee_2'])} | "
                f"regroup. {fmt_duree(cycle['duree_regroupement_2_min'])} | "
                f"maintien {fmt_duree(cycle['duree_maintien_2_min'])} | "
                f"sortie {fmt_duree(cycle['duree_sortie_2_min'])}")
    points_p2 = [
        (cycle['t4'], cycle['seuil_t4'], COULEURS_POINTS['t4'], LABELS_POINTS['t4'], True),
        (cycle['t5'], cycle['seuil_t5'], COULEURS_POINTS['t5'], LABELS_POINTS['t5'], True),
        (cycle['t6'], cycle['seuil_t6'], COULEURS_POINTS['t6'], LABELS_POINTS['t6'], True),
        (cycle['t7'], cycle['seuil_t7'], COULEURS_POINTS['t7'], LABELS_POINTS['t7'], True),
    ]
    if cycle['instant_maintien_2'] is not None and cycle['temp_maintien_2'] is not None:
        points_p2.append((cycle['instant_maintien_2'], cycle['temp_maintien_2'],
                           _COULEUR_MAX_MAINTIEN, "Max maintien 2", True))
    _tracer_zoom_phase(ax_p2, dates, voies, premiere_nom, derniere_nom, config, points_p2, titre_p2)

    # --- Phase 3 : refroidissement palier 3 (descente / regroupement) ---
    titre_p3 = (f"Refroidissement palier 3 — descente {fmt_vitesse_h(cycle['vitesse_descente_3'])} | "
                f"regroup. {fmt_duree(cycle['duree_regroupement_3_min'])}")
    points_p3 = [
        (cycle['t7'], cycle['seuil_t7'], COULEURS_POINTS['t7'], LABELS_POINTS['t7'], True),
        (cycle['t8'], cycle['seuil_t8'], COULEURS_POINTS['t8'], LABELS_POINTS['t8'], True),
        (cycle['t9'], cycle['seuil_t9'], COULEURS_POINTS['t9'], LABELS_POINTS['t9'], True),
    ]
    _tracer_zoom_phase(ax_p3, dates, voies, premiere_nom, derniere_nom, config, points_p3, titre_p3)

    # --- Phase 4 : sortie finale palier 4 ---
    titre_p4 = f"Sortie finale — descente {fmt_vitesse_h(cycle['vitesse_descente_4'])}"
    points_p4 = [
        (cycle['t9'], cycle['seuil_t9'], COULEURS_POINTS['t9'], LABELS_POINTS['t9'], True),
        (cycle['t10'], cycle['seuil_t10'], COULEURS_POINTS['t10'], LABELS_POINTS['t10'], True),
    ]
    _tracer_zoom_phase(ax_p4, dates, voies, premiere_nom, derniere_nom, config, points_p4, titre_p4)

    fig.suptitle(titre or os.path.basename(fichier), fontsize=14, fontweight='bold', y=0.99)
    fig.text(0.5, 0.965,
              f"1ère voie : {premiere_nom}   |   dernière voie : {derniere_nom}   |   "
              f"T° initiale : {fmt_temperature(cycle['temp_initiale'])}   |   "
              f"T° maintien 2 (max) : {fmt_temperature(cycle['temp_maintien_2'])}",
              ha='center', fontsize=9, color='dimgray')

    fig.subplots_adjust(top=0.90, bottom=0.14, left=0.06, right=0.95)
    return fig


# =========================================================================
# Fichier de résultats (formalisme strict, voir hypothèse 8)
# =========================================================================

# (position_tabulation, libellé) pour chacune des 12 valeurs, dans l'ordre.
# 'avant' -> "\tvaleur" ; 'apres' -> "valeur\t" (reproduit exactement le
# formalisme — et l'orthographe — de l'exemple fourni).
_FORMALISME_RESULTATS = [
    ('avant', "température initiale"),
    ('avant', "vitesse de montée 1"),
    ('avant', "durée de regroupement 1"),
    ('avant', "durée du maintient 1"),
    ('apres', "Vitesse de montée 2"),
    ('apres', "Durée de regroupement 2"),
    ('apres', "Température de maintient 2"),
    ('apres', "Durée du maintient 2"),
    ('apres', "Durée de la sortie 2"),
    ('apres', "Vitesse de descente 3"),
    ('avant', "Durée du regroupement 3"),
    ('avant', "vitese de descente 4"),
]


def ecrire_resultats(fichier_sortie, cycle):
    valeurs = [
        fmt_temperature(cycle['temp_initiale']),
        fmt_vitesse_h(cycle['vitesse_montee_1']),
        fmt_duree(cycle['duree_regroupement_1_min']),
        fmt_duree(cycle['duree_maintien_1_min']),
        fmt_vitesse_h(cycle['vitesse_montee_2']),
        fmt_duree(cycle['duree_regroupement_2_min']),
        fmt_temperature(cycle['temp_maintien_2']),
        fmt_duree(cycle['duree_maintien_2_min']),
        fmt_duree(cycle['duree_sortie_2_min']),
        fmt_vitesse_h(cycle['vitesse_descente_3']),
        fmt_duree(cycle['duree_regroupement_3_min']),
        fmt_vitesse_h(cycle['vitesse_descente_4']),
    ]

    lignes_libelles = []
    lignes_valeurs = []
    for (position, libelle), valeur in zip(_FORMALISME_RESULTATS, valeurs):
        if position == 'avant':
            lignes_libelles.append(f"\t{libelle}")
            lignes_valeurs.append(f"\t{valeur}")
        else:
            lignes_libelles.append(f"{libelle}\t")
            lignes_valeurs.append(f"{valeur}\t")

    with open(fichier_sortie, 'w', encoding='utf-8', newline='') as f:
        f.write("\n".join(lignes_libelles + lignes_valeurs) + "\n")


# =========================================================================
# Traitement d'un fichier
# =========================================================================

def analyser_fichier(fichier, config):
    print(f"📄 Lecture de {fichier}...")
    titre, dates, valeurs_brutes = lire_donnees_excel(fichier)
    print(f"   {len(dates)} échantillons, {len(valeurs_brutes)} voies détectées.")

    voies = filtrer_voies(valeurs_brutes, config['CH_ignore_dT'])
    voies_ignorees = sorted(set(valeurs_brutes) - set(voies))
    if voies_ignorees:
        print(f"   ⏭️  Voies ignorées (-OVER ou ΔT ≤ {config['CH_ignore_dT']}°) : "
              f"{', '.join(voies_ignorees)}")
    if not voies:
        print(f"   ⚠️ Aucune voie exploitable pour {fichier}, fichier ignoré.")
        return
    print(f"   ✅ Voies retenues : {', '.join(voies)}")

    cycle = calculer_cycle(dates, voies, config)
    print(f"   📐 1ère voie : {cycle['premiere_nom']}   dernière voie : {cycle['derniere_nom']}")
    print(f"   T° initiale = {fmt_temperature(cycle['temp_initiale'])}")

    etapes = [
        ('t0', 'début montée', "temp_initiale + temp_delta_ambiant"),
        ('t1', 'fin montée 1', "temp_palier_1 - temp_delta_1"),
        ('t2', 'regroupement 1', "temp_palier_1 - temp_delta_1 (dernière voie)"),
        ('t3', 'maintien 1', "temp_palier_1 + temp_delta_1"),
        ('t4', 'fin montée 2', "temp_palier_2 - temp_delta_2"),
        ('t5', 'regroupement 2', "temp_palier_2 - temp_delta_2 (dernière voie)"),
        ('t6', 'maintien 2', "temp_palier_2 + temp_delta_2"),
        ('t7', 'sortie 2', "temp_palier_2 - temp_delta_2 (dernière voie, descente)"),
        ('t8', 'fin descente 3', "temp_palier_3 + temp_delta_3"),
        ('t9', 'regroupement 3', "temp_palier_3 - temp_delta_3 (dernière voie)"),
        ('t10', 'fin descente 4', "temp_palier_4"),
    ]
    for cle, nom_etape, _ in etapes:
        if cycle[cle] is None:
            print(f"   ⚠️ Étape non atteinte : {nom_etape} ({cle}).")
            break
    else:
        print(f"   ⏱️  Cycle complet détecté (t0 → t10).")

    fig = tracer_graphique(fichier, titre, dates, voies, cycle, config)

    base, _ = os.path.splitext(fichier)
    fichier_image = f"{base}_{config['suffixe']}.{FORMAT_IMAGE}"
    fig.savefig(fichier_image, dpi=config['dpi'], bbox_inches='tight')
    plt.close(fig)
    print(f"✅ Graphique généré -> {fichier_image}")

    fichier_resultats = f"{base}_{config['suffixe_resultats']}.{FORMAT_RESULTATS}"
    ecrire_resultats(fichier_resultats, cycle)
    print(f"✅ Résultats générés -> {fichier_resultats}")


# =========================================================================
# Configuration (YAML) et arguments en ligne de commande
# =========================================================================

DEFAULTS = {
    'CH_ignore_dT': 20.0,
    'temp_delta_ambiant': 10.0,
    'temp_palier_1': 450.0,
    'temp_delta_1': 15.0,
    'temp_palier_2': 550.0,
    'temp_delta_2': 15.0,
    'temp_palier_3': 300.0,
    'temp_delta_3': 15.0,
    'temp_palier_4': 50.0,
    'zoom_marge_temperature': 20.0,
    'zoom_marge_temps': 5.0,
    'suffixe': 'Analyse',
    'suffixe_resultats': 'Resultats',
    'dpi': 150,
    'figure_taille': [16, 11],
}


def lire_parametres_yaml(fichier):
    if not os.path.exists(fichier):
        raise FileNotFoundError(f"Le fichier {fichier} est introuvable.")
    with open(fichier, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f) or {}


def construire_config(args):
    fichier_config = args.config or DEFAULT_CONFIG_FILE
    yaml_params = {}
    if os.path.exists(fichier_config):
        print(f"⚙️ Lecture des paramètres depuis {fichier_config}...")
        yaml_params = lire_parametres_yaml(fichier_config)
    elif args.config:
        raise FileNotFoundError(f"Le fichier {fichier_config} est introuvable.")
    else:
        print(f"⚙️ Aucun fichier {fichier_config} trouvé, utilisation des valeurs par défaut.")

    def valeur(cli_val, cle_yaml):
        if cli_val is not None:
            return cli_val
        return yaml_params.get(cle_yaml, DEFAULTS[cle_yaml])

    config = {
        'CH_ignore_dT': float(valeur(args.ch_ignore_dt, 'CH_ignore_dT')),
        'temp_delta_ambiant': float(valeur(args.temp_delta_ambiant, 'temp_delta_ambiant')),
        'temp_palier_1': float(valeur(args.temp_palier_1, 'temp_palier_1')),
        'temp_delta_1': float(valeur(args.temp_delta_1, 'temp_delta_1')),
        'temp_palier_2': float(valeur(args.temp_palier_2, 'temp_palier_2')),
        'temp_delta_2': float(valeur(args.temp_delta_2, 'temp_delta_2')),
        'temp_palier_3': float(valeur(args.temp_palier_3, 'temp_palier_3')),
        'temp_delta_3': float(valeur(args.temp_delta_3, 'temp_delta_3')),
        'temp_palier_4': float(valeur(args.temp_palier_4, 'temp_palier_4')),
        'zoom_marge_temperature': float(
            yaml_params.get('zoom_marge_temperature', DEFAULTS['zoom_marge_temperature'])
        ),
        'zoom_marge_temps': float(
            yaml_params.get('zoom_marge_temps', DEFAULTS['zoom_marge_temps'])
        ),
        'suffixe': str(yaml_params.get('suffixe', DEFAULTS['suffixe'])),
        'suffixe_resultats': str(yaml_params.get('suffixe_resultats', DEFAULTS['suffixe_resultats'])),
        'dpi': int(yaml_params.get('dpi', DEFAULTS['dpi'])),
        'figure_taille': yaml_params.get('figure_taille', DEFAULTS['figure_taille']),
    }

    return config


def main():
    parser = argparse.ArgumentParser(
        description="Analyse du cycle d'hypertrempe (montée / regroupement / maintien x2 + "
                    "refroidissement x2) d'un ou plusieurs fichiers Excel et génération "
                    "d'images + fichiers de résultats."
    )
    parser.add_argument("fichiers", nargs="*", help="Un ou plusieurs fichiers Excel (.xlsx) à traiter")
    parser.add_argument("--config", type=str, help="Chemin du fichier YAML de configuration")
    parser.add_argument("--ch-ignore-dt", dest="ch_ignore_dt", type=float,
                         help="Seuil d'amplitude (°C) en dessous duquel une voie est ignorée")
    parser.add_argument("--temp-delta-ambiant", dest="temp_delta_ambiant", type=float,
                         help="Écart (°C) au-dessus de la température initiale marquant le début de la montée")
    parser.add_argument("--temp-palier-1", dest="temp_palier_1", type=float, help="Température du palier 1")
    parser.add_argument("--temp-delta-1", dest="temp_delta_1", type=float, help="Demi-tolérance (°C) du palier 1")
    parser.add_argument("--temp-palier-2", dest="temp_palier_2", type=float, help="Température du palier 2")
    parser.add_argument("--temp-delta-2", dest="temp_delta_2", type=float, help="Demi-tolérance (°C) du palier 2")
    parser.add_argument("--temp-palier-3", dest="temp_palier_3", type=float, help="Température du palier 3 (refroidissement)")
    parser.add_argument("--temp-delta-3", dest="temp_delta_3", type=float, help="Demi-tolérance (°C) du palier 3")
    parser.add_argument("--temp-palier-4", dest="temp_palier_4", type=float, help="Température de sortie finale (palier 4)")

    args = parser.parse_args()

    try:
        config = construire_config(args)
    except (FileNotFoundError, ValueError) as e:
        print(f"❌ Erreur de configuration : {e}")
        sys.exit(1)

    print("=== Paramètres d'analyse ===")
    for cle in ('CH_ignore_dT', 'temp_delta_ambiant', 'temp_palier_1', 'temp_delta_1',
                'temp_palier_2', 'temp_delta_2', 'temp_palier_3', 'temp_delta_3',
                'temp_palier_4', 'zoom_marge_temperature', 'zoom_marge_temps',
                'suffixe', 'suffixe_resultats'):
        print(f"{cle} = {config[cle]}")
    print()

    if args.fichiers:
        print(f"🔍 {len(args.fichiers)} fichier(s) Excel fourni(s) :")
        for fichier in args.fichiers:
            try:
                analyser_fichier(fichier, config)
            except Exception as e:
                print(f"❌ Erreur lors du traitement de {fichier} : {e}")
    else:
        print("Traitement des fichiers .xlsx (hors fichiers déjà suffixés) du répertoire actuel.")
        suffixe_bas = f"_{config['suffixe']}".lower()
        suffixe_resultats_bas = f"_{config['suffixe_resultats']}".lower()
        fichiers_xlsx = [
            f for f in os.listdir('.')
            if f.lower().endswith('.xlsx') and not f.lower()[:-5].endswith(suffixe_bas)
            and not f.lower()[:-5].endswith(suffixe_resultats_bas)
        ]
        if not fichiers_xlsx:
            print("⚠️ Aucun fichier .xlsx trouvé dans le répertoire actuel.")
            return

        print(f"🔍 {len(fichiers_xlsx)} fichier(s) .xlsx trouvé(s) :")
        for f in fichiers_xlsx:
            print(f" → {f}")
        print("\n⏳ Traitement en cours...\n")

        for fichier in fichiers_xlsx:
            try:
                analyser_fichier(fichier, config)
            except Exception as e:
                print(f"❌ Erreur lors du traitement de {fichier} : {e}")

        print("\n✅ Tous les fichiers .xlsx ont été traités.")


if __name__ == "__main__":
    main()
