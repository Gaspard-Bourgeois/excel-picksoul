"""
analyseTemperatureExcel.py

Analyse un (ou plusieurs) fichier(s) Excel d'enregistreur de température
(type SMARTDAC+, colonnes CHxxxx / CHCxxx) et génère, pour chacun, une
image regroupant :
  - un graphique gauche (pleine hauteur) montrant l'ensemble des voies
    retenues après filtrage (voies grisées, courbe de référence surlignée),
  - un zoom en haut à droite sur la montée,
  - un zoom au centre à droite sur le maintien,
  - un zoom en bas à droite sur le refroidissement,
ainsi qu'un second fichier texte contenant les valeurs de synthèse du
cycle (température initiale, durée de montée, température maximale de
maintien, durée de maintien, vitesse de refroidissement), au format imposé
par un exemple fourni (voir hypothèse 8).

Les paramètres se trouvent dans un fichier YAML (par défaut
analyseTemperatureExcel.yaml), avec quelques surcharges possibles en ligne
de commande pour les seuils principaux.

Usage :
    python analyseTemperatureExcel.py fichier1.xlsx fichier2.xlsx
    (sans fichier -> traite tous les .xlsx du dossier courant, hors fichiers
    déjà suffixés)
    (paramètres absents de la ligne de commande -> lus dans
    analyseTemperatureExcel.yaml)

-------------------------------------------------------------------------
HYPOTHÈSES / CHOIX DE CONCEPTION (assumés faute de spécification exacte) :

1. Structure du fichier Excel : la ligne d'en-tête des voies est repérée
   automatiquement en cherchant la première ligne contenant des cellules du
   type "CH0001", "CH0002"... ou "CHC001", "CHC002"... (pas de numéro de
   ligne fixe). La ligne de début des données est repérée en cherchant,
   après cette ligne, la ligne dont la colonne A vaut "Date" et la colonne
   B vaut "Time" : les données commencent juste après. La lecture s'arrête
   à la première ligne dont la colonne "Date" est vide. Aucune correction
   de chronologie n'est appliquée : les dates/heures du fichier sont
   utilisées telles quelles.

2. Cellule `titre_graphique` : repérée en cherchant, en colonne A, la ligne
   dont le libellé vaut exactement "Batch No." (fixe, non paramétrable), et
   en lisant la valeur en colonne `titre_graphique_col` (par défaut colonne
   C, comme les autres champs "libellé / valeur" de cet en-tête).

3. `CH_ignore_dT` : une voie est CONSERVÉE si (max - min) de ses valeurs
   sur l'ensemble du fichier est STRICTEMENT SUPÉRIEUR à `CH_ignore_dT` ;
   sinon elle est ignorée (voie trop stable / non pertinente). Une voie
   contenant au moins une valeur de type "-OVER" (ou plus généralement
   toute chaîne contenant "OVER") est ignorée avant même ce calcul, quelle
   que soit son amplitude.

4. Courbe de référence (paramètre `courbe_reference`) : utilisée pour
   détecter les seuils de montée/maintien/refroidissement.
     - `"premiere"` (valeur par défaut) : la première voie NON filtrée,
       dans l'ordre des colonnes du fichier Excel.
     - `"moyenne"` : moyenne, à chaque instant, des voies retenues.
     - un nom de voie explicite (ex. `"CH0003"`).
   Toutes les AUTRES voies retenues sont tracées en gris (grisées, sans
   distinction de couleur entre elles) sur les 4 graphiques ; la courbe de
   référence est surlignée en noir, plus épaisse, avec son nom en légende.

5. Détection des instants (recherche séquentielle, chaque étape reprenant
   la recherche à partir de l'instant trouvé à l'étape précédente, pour
   éviter de détecter un franchissement antérieur non pertinent) :
     - t1 = premier instant où la courbe de référence dépasse (en montant)
       `temp_montee_basse`,
     - t2 = premier instant, après t1, où elle dépasse (en montant)
       `temp_montee_haute` -> durée de montée = t2 - t1,
     - t3 = premier instant, après t2, où elle dépasse (en montant)
       `temp_maintien_debut`,
     - t4 = premier instant, après t3, où elle repasse (en descendant)
       sous `temp_maintien_fin` -> durée de maintien = t4 - t3 ; la
       température maximale de maintien est le maximum de la courbe de
       référence entre t3 et t4 (inclus),
     - t5 = premier instant, après t4, où elle repasse (en descendant)
       sous `temp_refroidissement` (par défaut = `temp_montee_basse` si
       non renseigné) -> vitesse de refroidissement =
       (valeur en t4 - valeur en t5) / durée(t4->t5), en °C/min.
   Tous les franchissements sont interpolés linéairement entre les deux
   échantillons encadrants (précision sub-échantillon). Si une étape
   n'est pas trouvée, les étapes suivantes qui en dépendent sont
   également non calculées (valeurs "N/A" dans les sorties).

6. Fenêtres de zoom : deux paramètres uniques, appliqués aux 3 zooms
   (montée / maintien / refroidissement) :
     - `zoom_marge_temperature` (°C) : marge ajoutée de part et d'autre de
       la plage de températures utile de chaque zoom,
     - `zoom_marge_temps` (minutes) : marge ajoutée de part et d'autre de
       la plage temporelle utile de chaque zoom.

7. Fichier de sortie image : `<nom_fichier_excel_sans_extension>_<suffixe>.<format_image>`.
   Fichier de sortie résultats : `<nom_fichier_excel_sans_extension>_<suffixe_resultats>.<format_resultats>`.
   Les deux fichiers sont écrits dans le même dossier que le fichier Excel
   source.

8. Formalisme STRICT du fichier de résultats (imposé, reproduit tel quel) :
   chaque ligne de donnée comporte une valeur et une tabulation, la
   tabulation étant placée AVANT la valeur pour la 1ère ligne (température
   initiale) et APRÈS la valeur pour les 4 lignes suivantes, exactement
   comme dans l'exemple fourni :
       \t12°C
       5 min\t
       600°C\t
       5 min\t
       40°C/min\t
   Les libellés ("signification") sont ajoutés EN AMONT de ce bloc de
   données, avec le même formalisme (tabulation au même endroit que sur la
   ligne de donnée correspondante) :
       \ttempérature initiale
       durée de montée\t
       température maximale de maintien\t
       durée de maintien\t
       vitesse de refroidissement\t
   Le fichier final concatène donc le bloc des 5 libellés puis le bloc des
   5 valeurs, dans cet ordre : température initiale, durée de montée,
   température maximale de maintien, durée de maintien, vitesse de
   refroidissement.

9. Température initiale : première valeur (au premier instant du fichier)
   de la courbe de référence.

10. Toutes les valeurs de sortie (température initiale, durées, vitesse de
    refroidissement) sont arrondies à l'entier le plus proche, sans
    décimale (ex. "5 min", "39°C"). Formatage des durées en minutes plutôt
    qu'en "HH:MM:SS", le cycle thermique étudié se comptant en minutes.

11. Couleurs : chaque « point d'intérêt » (t1 montée basse, t2 montée
    haute, t3 maintien début, t4 maintien fin, t5 refroidissement) a sa
    propre couleur fixe, réutilisée à l'identique sur les 4 graphiques
    (trait de seuil, trait vertical, marqueur, texte d'annotation) plutôt
    qu'une couleur générique "début/fin" — utile notamment quand deux
    seuils ont la même valeur (ex. maintien_debut == maintien_fin) et que
    les traits se superposeraient sinon. Les annotations textuelles sont
    posées sur un fond blanc semi-opaque pour rester lisibles par-dessus
    les courbes sombres.
-------------------------------------------------------------------------
"""

import argparse
import os
import re
import sys
from datetime import datetime, timedelta

import matplotlib
matplotlib.use("Agg")
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import openpyxl
import yaml


# =========================================================================
# Lecture du fichier Excel
# =========================================================================

_RE_CHANNEL = re.compile(r'^CH[A-Z]?\d{3,4}$', re.IGNORECASE)
TITRE_GRAPHIQUE_LABEL = "Batch No."


def trouver_ligne_entete_channels(ws, max_rows_scan=200):
    """Cherche la première ligne contenant des libellés de voie CHxxxx /
    CHCxxx. Renvoie (numero_ligne, {colonne: nom_voie})."""
    for r in range(1, min(max_rows_scan, ws.max_row) + 1):
        canaux = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and _RE_CHANNEL.match(v.strip()):
                canaux[c] = v.strip()
        if canaux:
            return r, canaux
    raise ValueError("Impossible de trouver la ligne d'en-tête des voies (CHxxxx / CHCxxx).")


def trouver_ligne_debut_donnees(ws, ligne_apres, max_rows_scan=50):
    """Cherche, après la ligne d'en-tête des voies, la ligne 'Date'/'Time' et
    renvoie le numéro de la première ligne de données (juste après)."""
    borne = min(ligne_apres + max_rows_scan, ws.max_row)
    for r in range(ligne_apres, borne + 1):
        v1 = ws.cell(row=r, column=1).value
        v2 = ws.cell(row=r, column=2).value
        if isinstance(v1, str) and v1.strip().lower() == 'date' and \
                isinstance(v2, str) and v2.strip().lower() == 'time':
            return r + 1
    raise ValueError("Impossible de trouver la ligne d'en-tête 'Date' / 'Time'.")


def trouver_titre_graphique(ws, colonne, max_rows_scan=200):
    for r in range(1, min(max_rows_scan, ws.max_row) + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == TITRE_GRAPHIQUE_LABEL:
            val = ws.cell(row=r, column=colonne).value
            return "" if val is None else str(val).strip()
    return ""


def parser_date_heure(val_date, val_time):
    if isinstance(val_date, datetime):
        d = val_date.date()
    else:
        d = datetime.strptime(str(val_date).strip(), "%Y/%m/%d").date()

    if hasattr(val_time, 'hour'):
        t = val_time.time() if isinstance(val_time, datetime) else val_time
    else:
        s = str(val_time).strip()
        fmt = "%H:%M:%S" if s.count(':') == 2 else "%H:%M"
        t = datetime.strptime(s, fmt).time()

    return datetime.combine(d, t)


def lire_donnees_excel(fichier, config):
    wb = openpyxl.load_workbook(fichier, data_only=True)
    ws = wb[config['feuille']] if config['feuille'] else wb[wb.sheetnames[0]]

    ligne_ch, canaux_col = trouver_ligne_entete_channels(ws)
    ligne_debut = trouver_ligne_debut_donnees(ws, ligne_ch)
    titre = trouver_titre_graphique(ws, config['titre_graphique_col'])

    dates = []
    valeurs_brutes = {nom: [] for nom in canaux_col.values()}

    r = ligne_debut
    while True:
        v_date = ws.cell(row=r, column=1).value
        if v_date is None or str(v_date).strip() == '':
            break
        v_time = ws.cell(row=r, column=2).value
        dates.append(parser_date_heure(v_date, v_time))
        for c, nom in canaux_col.items():
            valeurs_brutes[nom].append(ws.cell(row=r, column=c).value)
        r += 1

    if not dates:
        raise ValueError(f"Aucune donnée trouvée dans {fichier}.")

    return titre, dates, valeurs_brutes


# =========================================================================
# Filtrage des voies
# =========================================================================

def _est_valeur_over(v):
    return isinstance(v, str) and 'OVER' in v.upper()


def filtrer_voies(valeurs_brutes, ch_ignore_dT):
    """Retire les voies contenant une valeur de type -OVER, puis celles dont
    l'amplitude (max-min) est <= ch_ignore_dT. Renvoie {nom: [floats]},
    en conservant l'ordre des colonnes du fichier Excel d'origine."""
    voies = {}
    for nom, valeurs in valeurs_brutes.items():
        if any(_est_valeur_over(v) for v in valeurs):
            continue
        try:
            floats = [float(v) for v in valeurs]
        except (TypeError, ValueError):
            continue
        if (max(floats) - min(floats)) > ch_ignore_dT:
            voies[nom] = floats
    return voies


def construire_courbe_reference(voies, courbe_reference):
    """Renvoie (nom_reference, [floats]) selon le paramètre courbe_reference :
    "premiere" (première voie non filtrée, dans l'ordre du fichier),
    "moyenne" (moyenne des voies retenues), ou un nom de voie explicite."""
    choix = courbe_reference.strip().lower()
    if choix == 'premiere':
        nom = next(iter(voies))
        return nom, voies[nom]
    if choix == 'moyenne':
        noms = list(voies)
        n = len(voies[noms[0]])
        moyenne = [sum(voies[nom][i] for nom in noms) / len(noms) for i in range(n)]
        return "Moyenne", moyenne
    if courbe_reference not in voies:
        raise ValueError(
            f"courbe_reference = {courbe_reference!r} : voie absente des voies retenues "
            f"({', '.join(sorted(voies))})."
        )
    return courbe_reference, voies[courbe_reference]


# =========================================================================
# Détection des instants de franchissement de seuil
# =========================================================================

def detecter_franchissement(dates, valeurs, seuil, sens, idx_debut=0):
    """sens = 'montant' (franchissement vers le haut) ou 'descente'
    (franchissement vers le bas). Recherche à partir de l'indice idx_debut.
    Renvoie (instant interpolé, indice de l'échantillon suivant le
    franchissement) ou (None, None) si le seuil n'est jamais franchi."""
    depart = max(1, idx_debut)
    for i in range(depart, len(valeurs)):
        v0, v1 = valeurs[i - 1], valeurs[i]
        if sens == 'montant':
            franchi = v0 < seuil <= v1
        else:
            franchi = v0 >= seuil > v1
        if not franchi:
            continue
        t0, t1 = dates[i - 1], dates[i]
        frac = 0.0 if v1 == v0 else (seuil - v0) / (v1 - v0)
        frac = max(0.0, min(1.0, frac))
        return t0 + (t1 - t0) * frac, i
    return None, None


def calculer_cycle(dates, ref, config):
    """Calcule les instants t1..t5 et les valeurs de synthèse du cycle
    thermique (voir hypothèse 5 en tête de fichier). Renvoie un dict."""
    resultat = {
        't1': None, 't2': None, 't3': None, 't4': None, 't5': None,
        'temp_initiale': ref[0] if ref else None,
        'duree_montee_min': None,
        'temp_max_maintien': None,
        'duree_maintien_min': None,
        'vitesse_refroidissement': None,
    }

    t1, i1 = detecter_franchissement(dates, ref, config['temp_montee_basse'], 'montant', 0)
    resultat['t1'] = t1
    if t1 is None:
        return resultat

    t2, i2 = detecter_franchissement(dates, ref, config['temp_montee_haute'], 'montant', i1)
    resultat['t2'] = t2
    if t2 is None:
        return resultat
    resultat['duree_montee_min'] = (t2 - t1).total_seconds() / 60.0

    t3, i3 = detecter_franchissement(dates, ref, config['temp_maintien_debut'], 'montant', i2)
    resultat['t3'] = t3
    if t3 is None:
        return resultat

    t4, i4 = detecter_franchissement(dates, ref, config['temp_maintien_fin'], 'descente', i3)
    resultat['t4'] = t4
    if t4 is None:
        return resultat
    resultat['duree_maintien_min'] = (t4 - t3).total_seconds() / 60.0
    resultat['temp_max_maintien'] = max(ref[i3:i4 + 1]) if i4 >= i3 else max(ref[i3], ref[i4])

    t5, i5 = detecter_franchissement(dates, ref, config['temp_refroidissement'], 'descente', i4)
    resultat['t5'] = t5
    if t5 is None:
        return resultat
    duree_refroidissement_min = (t5 - t4).total_seconds() / 60.0
    valeur_t4 = config['temp_maintien_fin']
    valeur_t5 = config['temp_refroidissement']
    if duree_refroidissement_min > 0:
        resultat['vitesse_refroidissement'] = (valeur_t4 - valeur_t5) / duree_refroidissement_min

    return resultat


# =========================================================================
# Formatage
# =========================================================================

def fmt_num(x, decimales=1):
    if x is None:
        return "N/A"
    s = f"{x:.{decimales}f}"
    if decimales > 0 and s.endswith('.' + '0' * decimales):
        s = s.split('.')[0]
    return s


def fmt_temperature(x):
    return "N/A" if x is None else f"{fmt_num(x, 0)}°C"


def fmt_duree_min(minutes):
    return "N/A" if minutes is None else f"{fmt_num(minutes, 0)} min"


def fmt_vitesse(x):
    return "N/A" if x is None else f"{fmt_num(x, 0)}°C/min"


# =========================================================================
# Graphique
# =========================================================================

_COULEUR_GRISE = '0.75'
_COULEUR_REFERENCE = 'black'

# Une couleur fixe par point d'intérêt, réutilisée à l'identique sur les
# 4 graphiques (voir hypothèse 11 en tête de fichier).
COULEURS_POINTS = {
    't0': '#8c564b',  # brun    - température initiale
    't1': '#1f77b4',  # bleu    - montée basse
    't2': '#ff7f0e',  # orange  - montée haute
    't3': '#2ca02c',  # vert    - maintien début
    't4': '#d62728',  # rouge   - maintien fin
    't5': '#9467bd',  # violet  - refroidissement
}
LABELS_POINTS = {
    't0': 'Température initiale',
    't1': 'Montée basse',
    't2': 'Montée haute',
    't3': 'Maintien début',
    't4': 'Maintien fin',
    't5': 'Refroidissement',
}
_COULEUR_MAX_MAINTIEN = 'teal'


def _tracer_courbes_fenetre(ax, dates, voies, nom_ref, ref, t_min, t_max):
    """Trace, dans la fenêtre [t_min, t_max], toutes les voies retenues en
    gris (une seule entrée de légende "Autres voies"), et la courbe de
    référence surlignée en noir."""
    premiere_grise = True
    for nom, valeurs in voies.items():
        if nom == nom_ref:
            continue
        xs, ys = [], []
        for t, v in zip(dates, valeurs):
            if t_min <= t <= t_max:
                xs.append(t)
                ys.append(v)
        if xs:
            label = "Autres voies" if premiere_grise else None
            ax.plot(xs, ys, color=_COULEUR_GRISE, linewidth=0.8, alpha=0.9,
                     zorder=1, label=label)
            premiere_grise = False

    xs_ref = [t for t in dates if t_min <= t <= t_max]
    ys_ref = [v for t, v in zip(dates, ref) if t_min <= t <= t_max]
    if xs_ref:
        ax.plot(xs_ref, ys_ref, color=_COULEUR_REFERENCE, linewidth=2, label=nom_ref, zorder=3)


def _annoter_point(ax, instant, valeur, texte, couleur, decalage=(6, 6), ha='left', label=None):
    ax.plot([instant], [valeur], marker='o', color=couleur, markersize=6,
             markeredgecolor='white', markeredgewidth=0.7, zorder=6, label=label)
    ax.annotate(texte, xy=(instant, valeur), xytext=decalage, textcoords='offset points',
                 color=couleur, fontsize=8, fontweight='bold', ha=ha,
                 bbox=dict(boxstyle='round,pad=0.2', facecolor='white',
                           edgecolor=couleur, linewidth=1, alpha=0.9))


def _tracer_zoom_montee(ax, dates, voies, nom_ref, ref, cycle, config):
    t1, t2 = cycle['t1'], cycle['t2']
    if t1 is None or t2 is None:
        ax.set_title("Montée : seuils non atteints")
        ax.text(0.5, 0.5, "Seuils non franchis", ha='center', va='center',
                 transform=ax.transAxes, color='dimgray')
        return

    marge_t = timedelta(minutes=config['zoom_marge_temps'])
    marge_T = config['zoom_marge_temperature']
    t_min, t_max = t1 - marge_t, t2 + marge_t
    y_min = config['temp_montee_basse'] - marge_T
    y_max = config['temp_montee_haute'] + marge_T

    _tracer_courbes_fenetre(ax, dates, voies, nom_ref, ref, t_min, t_max)
    ax.axhline(config['temp_montee_basse'], color=COULEURS_POINTS['t1'], linestyle='--', linewidth=1.2)
    ax.axhline(config['temp_montee_haute'], color=COULEURS_POINTS['t2'], linestyle='--', linewidth=1.2)
    _annoter_point(ax, t1, config['temp_montee_basse'],
                    f"{t1.strftime('%H:%M:%S')}\n{fmt_temperature(config['temp_montee_basse'])}",
                    COULEURS_POINTS['t1'])
    _annoter_point(ax, t2, config['temp_montee_haute'],
                    f"{t2.strftime('%H:%M:%S')}\n{fmt_temperature(config['temp_montee_haute'])}",
                    COULEURS_POINTS['t2'], decalage=(-8, 6), ha='right')

    ax.set_xlim(t_min, t_max)
    ax.set_ylim(y_min, y_max)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
    for lbl in ax.get_xticklabels():
        lbl.set_rotation(20)
        lbl.set_ha('right')
    ax.set_ylabel("Température (°C)")
    ax.grid(True, alpha=0.3)
    ax.set_title(f"Montée — durée : {fmt_duree_min(cycle['duree_montee_min'])}", fontweight='bold', pad=14)


def _tracer_zoom_maintien(ax, dates, voies, nom_ref, ref, cycle, config):
    t3, t4 = cycle['t3'], cycle['t4']
    if t3 is None or t4 is None:
        ax.set_title("Maintien : seuils non atteints")
        ax.text(0.5, 0.5, "Seuils non franchis", ha='center', va='center',
                 transform=ax.transAxes, color='dimgray')
        return

    marge_t = timedelta(minutes=config['zoom_marge_temps'])
    marge_T = config['zoom_marge_temperature']
    t_min, t_max = t3 - marge_t, t4 + marge_t
    t_max_maintien = cycle['temp_max_maintien']
    y_min = min(config['temp_maintien_debut'], config['temp_maintien_fin']) - marge_T
    y_max = t_max_maintien + marge_T

    _tracer_courbes_fenetre(ax, dates, voies, nom_ref, ref, t_min, t_max)
    ax.axhline(config['temp_maintien_debut'], color=COULEURS_POINTS['t3'], linestyle='--', linewidth=1.2)
    ax.axhline(config['temp_maintien_fin'], color=COULEURS_POINTS['t4'], linestyle='--', linewidth=1.2)
    ax.axhline(t_max_maintien, color=_COULEUR_MAX_MAINTIEN, linestyle=':', linewidth=1.3)
    _annoter_point(ax, t3, config['temp_maintien_debut'],
                    f"{t3.strftime('%H:%M:%S')}\n{fmt_temperature(config['temp_maintien_debut'])}",
                    COULEURS_POINTS['t3'])
    _annoter_point(ax, t4, config['temp_maintien_fin'],
                    f"{t4.strftime('%H:%M:%S')}\n{fmt_temperature(config['temp_maintien_fin'])}",
                    COULEURS_POINTS['t4'], decalage=(-8, 6), ha='right')
    ax.annotate(f"Max : {fmt_temperature(t_max_maintien)}", xy=(t_max, t_max_maintien),
                 xytext=(-6, 6), textcoords='offset points', ha='right',
                 color=_COULEUR_MAX_MAINTIEN, fontsize=8, fontweight='bold',
                 bbox=dict(boxstyle='round,pad=0.2', facecolor='white',
                           edgecolor=_COULEUR_MAX_MAINTIEN, linewidth=1, alpha=0.9))

    ax.set_xlim(t_min, t_max)
    ax.set_ylim(y_min, y_max)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
    for lbl in ax.get_xticklabels():
        lbl.set_rotation(20)
        lbl.set_ha('right')
    ax.set_ylabel("Température (°C)")
    ax.grid(True, alpha=0.3)
    ax.set_title(f"Maintien — durée : {fmt_duree_min(cycle['duree_maintien_min'])}", fontweight='bold', pad=14)


def _tracer_zoom_refroidissement(ax, dates, voies, nom_ref, ref, cycle, config):
    t4, t5 = cycle['t4'], cycle['t5']
    if t4 is None or t5 is None:
        ax.set_title("Refroidissement : seuils non atteints")
        ax.text(0.5, 0.5, "Seuils non franchis", ha='center', va='center',
                 transform=ax.transAxes, color='dimgray')
        return

    marge_t = timedelta(minutes=config['zoom_marge_temps'])
    marge_T = config['zoom_marge_temperature']
    t_min, t_max = t4 - marge_t, t5 + marge_t
    y_min = config['temp_refroidissement'] - marge_T
    y_max = config['temp_maintien_fin'] + marge_T

    _tracer_courbes_fenetre(ax, dates, voies, nom_ref, ref, t_min, t_max)
    ax.axhline(config['temp_maintien_fin'], color=COULEURS_POINTS['t4'], linestyle='--', linewidth=1.2)
    ax.axhline(config['temp_refroidissement'], color=COULEURS_POINTS['t5'], linestyle='--', linewidth=1.2)
    _annoter_point(ax, t4, config['temp_maintien_fin'],
                    f"{t4.strftime('%H:%M:%S')}\n{fmt_temperature(config['temp_maintien_fin'])}",
                    COULEURS_POINTS['t4'])
    _annoter_point(ax, t5, config['temp_refroidissement'],
                    f"{t5.strftime('%H:%M:%S')}\n{fmt_temperature(config['temp_refroidissement'])}",
                    COULEURS_POINTS['t5'], decalage=(-8, 6), ha='right')

    ax.set_xlim(t_min, t_max)
    ax.set_ylim(y_min, y_max)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
    for lbl in ax.get_xticklabels():
        lbl.set_rotation(20)
        lbl.set_ha('right')
    ax.set_ylabel("Température (°C)")
    ax.grid(True, alpha=0.3)
    ax.set_title(f"Refroidissement — vitesse : {fmt_vitesse(cycle['vitesse_refroidissement'])}",
                  fontweight='bold', pad=14)


def tracer_graphique(fichier, titre, dates, voies, nom_ref, ref, cycle, config):
    fig = plt.figure(figsize=tuple(config['figure_taille']))
    gs = fig.add_gridspec(3, 2, width_ratios=[1.4, 1], hspace=0.6, wspace=0.25)
    ax_gauche = fig.add_subplot(gs[:, 0])
    ax_montee = fig.add_subplot(gs[0, 1])
    ax_maintien = fig.add_subplot(gs[1, 1])
    ax_refroidissement = fig.add_subplot(gs[2, 1])

    # --- Graphique global (gauche, pleine hauteur) : voies grisées + référence ---
    premiere_grise = True
    for nom, valeurs in voies.items():
        if nom == nom_ref:
            continue
        label = "Autres voies" if premiere_grise else None
        ax_gauche.plot(dates, valeurs, color=_COULEUR_GRISE, linewidth=0.8,
                        alpha=0.9, label=label, zorder=1)
        premiere_grise = False
    ax_gauche.plot(dates, ref, color=_COULEUR_REFERENCE, linewidth=2, label=nom_ref, zorder=3)

    if cycle['temp_initiale'] is not None and dates:
        _annoter_point(ax_gauche, dates[0], cycle['temp_initiale'],
                        f"{dates[0].strftime('%H:%M:%S')}\n{fmt_temperature(cycle['temp_initiale'])}",
                        COULEURS_POINTS['t0'], decalage=(8, 8), label=LABELS_POINTS['t0'])

    for cle, couleur in COULEURS_POINTS.items():
        if cle == 't0':
            continue
        instant = cycle[cle]
        if instant is not None:
            ax_gauche.axvline(instant, color=couleur, linestyle='--', linewidth=1.3,
                                alpha=0.8, label=LABELS_POINTS[cle])
    ax_gauche.set_xlabel("Temps")
    ax_gauche.set_ylabel("Température (°C)")
    ax_gauche.xaxis.set_major_formatter(mdates.DateFormatter('%Y/%m/%d %H:%M:%S'))
    for lbl in ax_gauche.get_xticklabels():
        lbl.set_rotation(30)
        lbl.set_ha('right')
    ax_gauche.grid(True, alpha=0.3)
    ax_gauche.legend(loc='upper left', ncol=2, fontsize=7, framealpha=0.9)

    # --- Zooms montée / maintien / refroidissement ---
    _tracer_zoom_montee(ax_montee, dates, voies, nom_ref, ref, cycle, config)
    _tracer_zoom_maintien(ax_maintien, dates, voies, nom_ref, ref, cycle, config)
    _tracer_zoom_refroidissement(ax_refroidissement, dates, voies, nom_ref, ref, cycle, config)

    fig.suptitle(titre or os.path.basename(fichier), fontsize=14, fontweight='bold', y=0.99)
    fig.text(0.5, 0.955,
              f"Référence : {nom_ref}   |   "
              f"T° initiale : {fmt_temperature(cycle['temp_initiale'])}   |   "
              f"Durée montée : {fmt_duree_min(cycle['duree_montee_min'])}   |   "
              f"T° max maintien : {fmt_temperature(cycle['temp_max_maintien'])}   |   "
              f"Durée maintien : {fmt_duree_min(cycle['duree_maintien_min'])}   |   "
              f"Vitesse refroid. : {fmt_vitesse(cycle['vitesse_refroidissement'])}",
              ha='center', fontsize=9, color='dimgray')

    fig.subplots_adjust(top=0.88, bottom=0.16, left=0.06, right=0.95)
    return fig


# =========================================================================
# Fichier de résultats (formalisme strict, voir hypothèse 8)
# =========================================================================

# (position_tabulation, libellé) pour chacune des 5 valeurs, dans l'ordre.
# 'avant' -> "\tvaleur" ; 'apres' -> "valeur\t" (reproduit exactement le
# formalisme de l'exemple fourni).
_FORMALISME_RESULTATS = [
    ('avant', "température initiale"),
    ('apres', "durée de montée"),
    ('apres', "température maximale de maintien"),
    ('apres', "durée de maintien"),
    ('apres', "vitesse de refroidissement"),
]


def ecrire_resultats(fichier_sortie, cycle):
    valeurs = [
        fmt_temperature(cycle['temp_initiale']),
        fmt_duree_min(cycle['duree_montee_min']),
        fmt_temperature(cycle['temp_max_maintien']),
        fmt_duree_min(cycle['duree_maintien_min']),
        fmt_vitesse(cycle['vitesse_refroidissement']),
    ]

    lignes_libelles = []
    lignes_valeurs = []
    for (position, libelle), valeur in zip(_FORMALISME_RESULTATS, valeurs):
        if position == 'avant':
            lignes_libelles.append(f"\t{libelle}")
            lignes_valeurs.append(f"\t{valeur}")
        else:
            lignes_libelles.append(f"{libelle}\t")
            lignes_valeurs.append(f"{valeur}\t")

    with open(fichier_sortie, 'w', encoding='utf-8', newline='') as f:
        f.write("\n".join(lignes_libelles + lignes_valeurs) + "\n")


# =========================================================================
# Traitement d'un fichier
# =========================================================================

def analyser_fichier(fichier, config):
    print(f"📄 Lecture de {fichier}...")
    titre, dates, valeurs_brutes = lire_donnees_excel(fichier, config)
    print(f"   {len(dates)} échantillons, {len(valeurs_brutes)} voies détectées.")

    voies = filtrer_voies(valeurs_brutes, config['CH_ignore_dT'])
    voies_ignorees = sorted(set(valeurs_brutes) - set(voies))
    if voies_ignorees:
        print(f"   ⏭️  Voies ignorées (-OVER ou ΔT ≤ {config['CH_ignore_dT']}°) : "
              f"{', '.join(voies_ignorees)}")
    if not voies:
        print(f"   ⚠️ Aucune voie exploitable pour {fichier}, fichier ignoré.")
        return
    print(f"   ✅ Voies retenues : {', '.join(voies)}")

    nom_ref, ref = construire_courbe_reference(voies, config['courbe_reference'])
    print(f"   📐 Courbe de référence : {nom_ref}")

    cycle = calculer_cycle(dates, ref, config)
    print(f"   T° initiale = {fmt_temperature(cycle['temp_initiale'])}")
    if cycle['t1'] is None:
        print(f"   ⚠️ temp_montee_basse ({config['temp_montee_basse']}°) jamais franchie.")
    elif cycle['t2'] is None:
        print(f"   ⚠️ temp_montee_haute ({config['temp_montee_haute']}°) jamais franchie.")
    else:
        print(f"   ⏱️  Montée : {cycle['t1']} → {cycle['t2']} "
              f"(durée = {fmt_duree_min(cycle['duree_montee_min'])})")
        if cycle['t3'] is None:
            print(f"   ⚠️ temp_maintien_debut ({config['temp_maintien_debut']}°) jamais franchie.")
        elif cycle['t4'] is None:
            print(f"   ⚠️ temp_maintien_fin ({config['temp_maintien_fin']}°) jamais franchie.")
        else:
            print(f"   ⏱️  Maintien : {cycle['t3']} → {cycle['t4']} "
                  f"(durée = {fmt_duree_min(cycle['duree_maintien_min'])}, "
                  f"T° max = {fmt_temperature(cycle['temp_max_maintien'])})")
            if cycle['t5'] is None:
                print(f"   ⚠️ temp_refroidissement ({config['temp_refroidissement']}°) jamais franchie.")
            else:
                print(f"   ⏱️  Refroidissement : {cycle['t4']} → {cycle['t5']} "
                      f"(vitesse = {fmt_vitesse(cycle['vitesse_refroidissement'])})")

    fig = tracer_graphique(fichier, titre, dates, voies, nom_ref, ref, cycle, config)

    base, _ = os.path.splitext(fichier)
    fichier_image = f"{base}_{config['suffixe']}.{config['format_image']}"
    fig.savefig(fichier_image, dpi=config['dpi'], bbox_inches='tight')
    plt.close(fig)
    print(f"✅ Graphique généré -> {fichier_image}")

    fichier_resultats = f"{base}_{config['suffixe_resultats']}.{config['format_resultats']}"
    ecrire_resultats(fichier_resultats, cycle)
    print(f"✅ Résultats générés -> {fichier_resultats}")


# =========================================================================
# Configuration (YAML) et arguments en ligne de commande
# =========================================================================

DEFAULT_CONFIG_FILE = "analyseTemperatureExcel.yaml"

DEFAULTS = {
    'CH_ignore_dT': 20.0,
    'courbe_reference': 'premiere',
    'temp_montee_basse': 100.0,
    'temp_montee_haute': 200.0,
    'temp_maintien_debut': 200.0,
    'temp_maintien_fin': 200.0,
    'temp_refroidissement': None,  # None -> = temp_montee_haute
    'zoom_marge_temperature': 20.0,
    'zoom_marge_temps': 5.0,
    'suffixe': 'Analyse',
    'suffixe_resultats': 'Resultats',
    'format_image': 'png',
    'format_resultats': 'txt',
    'dpi': 150,
    'figure_taille': [15, 8],
    'feuille': None,
    'titre_graphique_col': 3,
}


def lire_parametres_yaml(fichier):
    if not os.path.exists(fichier):
        raise FileNotFoundError(f"Le fichier {fichier} est introuvable.")
    with open(fichier, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f) or {}


def construire_config(args):
    fichier_config = args.config or DEFAULT_CONFIG_FILE
    yaml_params = {}
    if os.path.exists(fichier_config):
        print(f"⚙️ Lecture des paramètres depuis {fichier_config}...")
        yaml_params = lire_parametres_yaml(fichier_config)
    elif args.config:
        raise FileNotFoundError(f"Le fichier {fichier_config} est introuvable.")
    else:
        print(f"⚙️ Aucun fichier {fichier_config} trouvé, utilisation des valeurs par défaut.")

    def valeur(cli_val, cle_yaml):
        if cli_val is not None:
            return cli_val
        return yaml_params.get(cle_yaml, DEFAULTS[cle_yaml])

    temp_montee_basse = float(valeur(args.temp_montee_basse, 'temp_montee_basse'))
    temp_montee_haute = float(valeur(args.temp_montee_haute, 'temp_montee_haute'))
    temp_refroidissement = valeur(args.temp_refroidissement, 'temp_refroidissement')
    temp_refroidissement = temp_montee_basse if temp_refroidissement is None else float(temp_refroidissement)

    config = {
        'CH_ignore_dT': float(valeur(args.ch_ignore_dt, 'CH_ignore_dT')),
        'courbe_reference': str(yaml_params.get('courbe_reference', DEFAULTS['courbe_reference'])),
        'temp_montee_basse': temp_montee_basse,
        'temp_montee_haute': temp_montee_haute,
        'temp_maintien_debut': float(valeur(args.temp_maintien_debut, 'temp_maintien_debut')),
        'temp_maintien_fin': float(valeur(args.temp_maintien_fin, 'temp_maintien_fin')),
        'temp_refroidissement': temp_refroidissement,
        'zoom_marge_temperature': float(
            yaml_params.get('zoom_marge_temperature', DEFAULTS['zoom_marge_temperature'])
        ),
        'zoom_marge_temps': float(
            yaml_params.get('zoom_marge_temps', DEFAULTS['zoom_marge_temps'])
        ),
        'suffixe': str(yaml_params.get('suffixe', DEFAULTS['suffixe'])),
        'suffixe_resultats': str(yaml_params.get('suffixe_resultats', DEFAULTS['suffixe_resultats'])),
        'format_image': str(yaml_params.get('format_image', DEFAULTS['format_image'])),
        'format_resultats': str(yaml_params.get('format_resultats', DEFAULTS['format_resultats'])),
        'dpi': int(yaml_params.get('dpi', DEFAULTS['dpi'])),
        'figure_taille': yaml_params.get('figure_taille', DEFAULTS['figure_taille']),
        'feuille': yaml_params.get('feuille', DEFAULTS['feuille']),
        'titre_graphique_col': int(
            yaml_params.get('titre_graphique_col', DEFAULTS['titre_graphique_col'])
        ),
    }

    return config


def main():
    parser = argparse.ArgumentParser(
        description="Analyse du cycle thermique (montée / maintien / refroidissement) d'un ou "
                    "plusieurs fichiers Excel et génération d'images + fichiers de résultats."
    )
    parser.add_argument("fichiers", nargs="*", help="Un ou plusieurs fichiers Excel (.xlsx) à traiter")
    parser.add_argument("--config", type=str, help="Chemin du fichier YAML de configuration")
    parser.add_argument("--ch-ignore-dt", dest="ch_ignore_dt", type=float,
                         help="Seuil d'amplitude (°C) en dessous duquel une voie est ignorée")
    parser.add_argument("--temp-montee-basse", dest="temp_montee_basse", type=float,
                         help="Température basse du seuil de montée")
    parser.add_argument("--temp-montee-haute", dest="temp_montee_haute", type=float,
                         help="Température haute du seuil de montée")
    parser.add_argument("--temp-maintien-debut", dest="temp_maintien_debut", type=float,
                         help="Température de début de maintien")
    parser.add_argument("--temp-maintien-fin", dest="temp_maintien_fin", type=float,
                         help="Température de fin de maintien")
    parser.add_argument("--temp-refroidissement", dest="temp_refroidissement", type=float,
                         help="Température cible de refroidissement (par défaut = temp_montee_haute)")

    args = parser.parse_args()

    try:
        config = construire_config(args)
    except (FileNotFoundError, ValueError) as e:
        print(f"❌ Erreur de configuration : {e}")
        sys.exit(1)

    print("=== Paramètres d'analyse ===")
    for cle in ('CH_ignore_dT', 'courbe_reference', 'temp_montee_basse', 'temp_montee_haute',
                'temp_maintien_debut', 'temp_maintien_fin', 'temp_refroidissement',
                'zoom_marge_temperature', 'zoom_marge_temps', 'suffixe', 'suffixe_resultats'):
        print(f"{cle} = {config[cle]}")
    print()

    if args.fichiers:
        print(f"🔍 {len(args.fichiers)} fichier(s) Excel fourni(s) :")
        for fichier in args.fichiers:
            try:
                analyser_fichier(fichier, config)
            except Exception as e:
                print(f"❌ Erreur lors du traitement de {fichier} : {e}")
    else:
        print("Traitement des fichiers .xlsx (hors fichiers déjà suffixés) du répertoire actuel.")
        suffixe_bas = f"_{config['suffixe']}".lower()
        suffixe_resultats_bas = f"_{config['suffixe_resultats']}".lower()
        fichiers_xlsx = [
            f for f in os.listdir('.')
            if f.lower().endswith('.xlsx') and not f.lower()[:-5].endswith(suffixe_bas)
            and not f.lower()[:-5].endswith(suffixe_resultats_bas)
        ]
        if not fichiers_xlsx:
            print("⚠️ Aucun fichier .xlsx trouvé dans le répertoire actuel.")
            return

        print(f"🔍 {len(fichiers_xlsx)} fichier(s) .xlsx trouvé(s) :")
        for f in fichiers_xlsx:
            print(f" → {f}")
        print("\n⏳ Traitement en cours...\n")

        for fichier in fichiers_xlsx:
            try:
                analyser_fichier(fichier, config)
            except Exception as e:
                print(f"❌ Erreur lors du traitement de {fichier} : {e}")

        print("\n✅ Tous les fichiers .xlsx ont été traités.")


if __name__ == "__main__":
    main()
