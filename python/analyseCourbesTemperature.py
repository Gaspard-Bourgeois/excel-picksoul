"""
analyseTemperatureExcel.py

Analyse un (ou plusieurs) fichier(s) Excel d'enregistreur de température
(type SMARTDAC+, colonnes CHxxxx) et génère, pour chacun, une image
regroupant :
  - un graphique bas (pleine largeur) montrant l'ensemble des voies
    retenues après filtrage,
  - un zoom en haut à gauche sur l'instant de franchissement du seuil de
    montée,
  - un zoom en haut à droite sur l'instant de franchissement du seuil de
    descente.

La logique de récupération des paramètres via un fichier YAML s'inspire de
decalageParachevementJBI.py : les paramètres se trouvent dans un fichier
YAML (par défaut analyseTemperatureExcel.yaml), avec quelques surcharges
possibles en ligne de commande pour les seuils principaux.

Usage :
    python analyseTemperatureExcel.py fichier1.xlsx fichier2.xlsx
    (sans fichier -> traite tous les .xlsx du dossier courant, hors fichiers
    déjà suffixés)
    (paramètres absents de la ligne de commande -> lus dans
    analyseTemperatureExcel.yaml)

-------------------------------------------------------------------------
HYPOTHÈSES / CHOIX DE CONCEPTION (assumés faute de spécification exacte) :

1. Structure du fichier Excel : la ligne d'en-tête des voies est repérée
   automatiquement en cherchant la première ligne contenant des cellules du
   type "CH0001", "CH0002"... (pas de numéro de ligne fixe). La ligne de
   début des données est repérée en cherchant, après cette ligne, la ligne
   dont la colonne A vaut "Date" et la colonne B vaut "Time" : les données
   commencent juste après. La lecture s'arrête à la première ligne dont la
   colonne "Date" est vide.

2. Cellule `titre_graphique` : repérée en cherchant, en colonne A, la ligne
   dont le libellé correspond au paramètre `titre_graphique_label` (par
   défaut "Batch No.", conforme au fichier d'exemple fourni), et en lisant
   la valeur en colonne `titre_graphique_col` (par défaut colonne C, comme
   les autres champs "libellé / valeur" de cet en-tête).

3. `CH_ignore_dT` : une voie est CONSERVÉE si (max - min) de ses valeurs
   (sur l'ensemble du fichier, avant rognage temporel) est STRICTEMENT
   SUPÉRIEUR à `CH_ignore_dT` ; sinon elle est ignorée (voie trop stable /
   non pertinente pour l'analyse thermique). Une voie contenant au moins
   une valeur de type "-OVER" (ou plus généralement toute chaîne contenant
   "OVER") est ignorée avant même ce calcul, quelle que soit son amplitude.

4. `time_ignore_dT` : pour chaque voie conservée, on cherche le premier
   instant où la valeur s'écarte de plus de `time_ignore_dT` de sa valeur
   de départ (respectivement de sa valeur d'arrivée, en partant de la fin).
   La fenêtre temporelle globale retenue est [début, fin] avec :
     - début = le PLUS PETIT de ces indices de départ parmi toutes les
       voies conservées (on rogne uniquement la portion initiale strictement
       plate pour TOUTES les voies),
     - fin = le PLUS GRAND des indices d'arrivée symétriques (on rogne
       uniquement la portion finale strictement plate pour TOUTES les
       voies).
   Si une voie ne varie jamais de plus de `time_ignore_dT`, elle ne
   contribue pas à réduire la fenêtre (bornes d'origine conservées pour
   elle).

5. Franchissement de seuil (montée / descente) : recherché par
   interpolation linéaire entre les deux échantillons encadrant le
   franchissement, pour un instant précis (et pas seulement à la
   résolution de l'échantillonnage). Pour une voie qui ne franchit jamais
   le seuil dans le sens demandé, elle est simplement exclue du calcul de
   l'instant de référence (elle reste toutefois affichée sur les
   graphiques).

6. `seuil_montant_first_or_last` / `seuil_descente_first_or_last` :
   valeur "first" -> on retient l'instant le plus TÔT parmi toutes les
   voies qui franchissent le seuil ; valeur "last" -> on retient l'instant
   le plus TARDIF (donc l'instant où TOUTES les voies concernées ont déjà
   franchi le seuil).

7. Fenêtres de zoom : `graphique_montant_dT` / `graphique_descente_dT`
   sont des demi-amplitudes en °C (axe Y centré sur le seuil, de
   seuil-dT à seuil+dT) ; `graphique_montant_dt` / `graphique_descente_dt`
   sont des demi-fenêtres temporelles en SECONDES (axe X centré sur
   l'instant détecté).

8. Fichier de sortie : `<nom_fichier_excel_sans_extension>_<suffixe>.<format_image>`,
   dans le même dossier que le fichier Excel source.
-------------------------------------------------------------------------
"""

import argparse
import os
import re
import sys
from datetime import datetime, timedelta

import matplotlib
matplotlib.use("Agg")
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import openpyxl
import yaml


# =========================================================================
# Lecture du fichier Excel
# =========================================================================

_RE_CHANNEL = re.compile(r'^CH\d{3,4}$', re.IGNORECASE)


def trouver_ligne_entete_channels(ws, max_rows_scan=200):
    """Cherche la première ligne contenant des libellés de voie CHxxxx.
    Renvoie (numero_ligne, {colonne: nom_voie})."""
    for r in range(1, min(max_rows_scan, ws.max_row) + 1):
        canaux = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and _RE_CHANNEL.match(v.strip()):
                canaux[c] = v.strip()
        if canaux:
            return r, canaux
    raise ValueError("Impossible de trouver la ligne d'en-tête des voies (CHxxxx).")


def trouver_ligne_debut_donnees(ws, ligne_apres, max_rows_scan=50):
    """Cherche, après la ligne d'en-tête des voies, la ligne 'Date'/'Time' et
    renvoie le numéro de la première ligne de données (juste après)."""
    borne = min(ligne_apres + max_rows_scan, ws.max_row)
    for r in range(ligne_apres, borne + 1):
        v1 = ws.cell(row=r, column=1).value
        v2 = ws.cell(row=r, column=2).value
        if isinstance(v1, str) and v1.strip().lower() == 'date' and \
                isinstance(v2, str) and v2.strip().lower() == 'time':
            return r + 1
    raise ValueError("Impossible de trouver la ligne d'en-tête 'Date' / 'Time'.")


def trouver_titre_graphique(ws, label, colonne, max_rows_scan=200):
    for r in range(1, min(max_rows_scan, ws.max_row) + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip() == label:
            val = ws.cell(row=r, column=colonne).value
            return "" if val is None else str(val).strip()
    return ""


def parser_date_heure(val_date, val_time):
    if isinstance(val_date, datetime):
        d = val_date.date()
    else:
        d = datetime.strptime(str(val_date).strip(), "%Y/%m/%d").date()

    if hasattr(val_time, 'hour'):
        t = val_time.time() if isinstance(val_time, datetime) else val_time
    else:
        s = str(val_time).strip()
        fmt = "%H:%M:%S" if s.count(':') == 2 else "%H:%M"
        t = datetime.strptime(s, fmt).time()

    return datetime.combine(d, t)


def lire_donnees_excel(fichier, config):
    wb = openpyxl.load_workbook(fichier, data_only=True)
    ws = wb[config['feuille']] if config['feuille'] else wb[wb.sheetnames[0]]

    ligne_ch, canaux_col = trouver_ligne_entete_channels(ws)
    ligne_debut = trouver_ligne_debut_donnees(ws, ligne_ch)
    titre = trouver_titre_graphique(
        ws, config['titre_graphique_label'], config['titre_graphique_col']
    )

    dates = []
    valeurs_brutes = {nom: [] for nom in canaux_col.values()}

    r = ligne_debut
    while True:
        v_date = ws.cell(row=r, column=1).value
        if v_date is None or str(v_date).strip() == '':
            break
        v_time = ws.cell(row=r, column=2).value
        dates.append(parser_date_heure(v_date, v_time))
        for c, nom in canaux_col.items():
            valeurs_brutes[nom].append(ws.cell(row=r, column=c).value)
        r += 1

    if not dates:
        raise ValueError(f"Aucune donnée trouvée dans {fichier}.")

    return titre, dates, valeurs_brutes


# =========================================================================
# Filtrage des voies et rognage temporel
# =========================================================================

def _est_valeur_over(v):
    return isinstance(v, str) and 'OVER' in v.upper()


def filtrer_voies(valeurs_brutes, ch_ignore_dT):
    """Retire les voies contenant une valeur de type -OVER, puis celles dont
    l'amplitude (max-min) est <= ch_ignore_dT. Renvoie {nom: [floats]}."""
    voies = {}
    for nom, valeurs in valeurs_brutes.items():
        if any(_est_valeur_over(v) for v in valeurs):
            continue
        try:
            floats = [float(v) for v in valeurs]
        except (TypeError, ValueError):
            continue
        if (max(floats) - min(floats)) > ch_ignore_dT:
            voies[nom] = floats
    return voies


def rogner_debut_fin(dates, voies, time_ignore_dT):
    """Rogne le début/la fin communs des courbes selon time_ignore_dT
    (voir hypothèse 4 en tête de fichier)."""
    n = len(dates)
    indices_debut = []
    indices_fin = []

    for valeurs in voies.values():
        v0 = valeurs[0]
        idx_d = n - 1
        for i, v in enumerate(valeurs):
            if abs(v - v0) > time_ignore_dT:
                idx_d = i
                break
        indices_debut.append(idx_d)

        vf = valeurs[-1]
        idx_f = 0
        for i in range(n - 1, -1, -1):
            if abs(valeurs[i] - vf) > time_ignore_dT:
                idx_f = i
                break
        indices_fin.append(idx_f)

    debut = min(indices_debut) if indices_debut else 0
    fin = max(indices_fin) if indices_fin else n - 1
    if fin <= debut:
        debut, fin = 0, n - 1

    dates_r = dates[debut:fin + 1]
    voies_r = {nom: v[debut:fin + 1] for nom, v in voies.items()}
    return dates_r, voies_r


# =========================================================================
# Détection des instants de franchissement de seuil
# =========================================================================

def detecter_franchissement(dates, valeurs, seuil, sens):
    """sens = 'montant' (franchissement vers le haut) ou 'descente'
    (franchissement vers le bas). Renvoie l'instant interpolé du premier
    franchissement, ou None si le seuil n'est jamais franchi dans ce sens."""
    for i in range(1, len(valeurs)):
        v0, v1 = valeurs[i - 1], valeurs[i]
        if sens == 'montant':
            franchi = v0 < seuil <= v1
        else:
            franchi = v0 >= seuil > v1
        if not franchi:
            continue
        t0, t1 = dates[i - 1], dates[i]
        frac = 0.0 if v1 == v0 else (seuil - v0) / (v1 - v0)
        frac = max(0.0, min(1.0, frac))
        return t0 + (t1 - t0) * frac
    return None


def instant_seuil(dates, voies, seuil, sens, first_or_last):
    instants = []
    for nom, valeurs in voies.items():
        t = detecter_franchissement(dates, valeurs, seuil, sens)
        if t is not None:
            instants.append((nom, t))
    if not instants:
        return None, []
    if first_or_last == 'first':
        nom_ref, instant = min(instants, key=lambda x: x[1])
    else:
        nom_ref, instant = max(instants, key=lambda x: x[1])
    return instant, instants


def formater_duree(td):
    total = int(td.total_seconds())
    signe = '-' if total < 0 else ''
    h, reste = divmod(abs(total), 3600)
    m, s = divmod(reste, 60)
    return f"{signe}{h:02d}:{m:02d}:{s:02d}"


# =========================================================================
# Graphique
# =========================================================================

def construire_couleurs(noms):
    noms = sorted(noms)
    n = len(noms)
    cmap = plt.get_cmap('tab20' if n <= 20 else 'nipy_spectral')
    if n <= 1:
        return {noms[0]: cmap(0.0)} if noms else {}
    return {nom: cmap(i / (n - 1)) for i, nom in enumerate(noms)}


def _tracer_zoom(ax, dates, voies, couleurs, instant, seuil, d_T, d_t, titre_zoom):
    if instant is None:
        ax.set_title(f"{titre_zoom} : seuil non atteint")
        ax.text(0.5, 0.5, "Seuil non franchi", ha='center', va='center',
                 transform=ax.transAxes, color='dimgray')
        ax.set_xticks([])
        ax.set_yticks([])
        return

    t_min = instant - timedelta(seconds=d_t)
    t_max = instant + timedelta(seconds=d_t)

    for nom, valeurs in sorted(voies.items()):
        xs = [t for t in dates if t_min <= t <= t_max]
        if not xs:
            continue
        ys = [v for t, v in zip(dates, valeurs) if t_min <= t <= t_max]
        ax.plot(xs, ys, color=couleurs[nom], marker='o', markersize=2.5,
                 linewidth=1, label=nom)

    ax.axhline(seuil, color='red', linestyle='--', linewidth=1.3, zorder=5)
    ax.axvline(instant, color='red', linestyle='--', linewidth=1.3, zorder=5)

    ax.set_ylim(seuil - d_T, seuil + d_T)
    ax.set_xlim(t_min, t_max)
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
    for lbl in ax.get_xticklabels():
        lbl.set_rotation(30)
        lbl.set_ha('right')
    ax.set_ylabel("Température (°C)")
    ax.grid(True, alpha=0.3)
    ax.set_title(titre_zoom, fontweight='bold')

    ax.annotate(
        f"{seuil:.1f} °C", xy=(1, seuil), xycoords=('axes fraction', 'data'),
        xytext=(-6, 4), textcoords='offset points', ha='right',
        color='red', fontsize=9, fontweight='bold'
    )
    ax.annotate(
        instant.strftime('%H:%M:%S'), xy=(instant, 1), xycoords=('data', 'axes fraction'),
        xytext=(6, -4), textcoords='offset points', va='top',
        color='red', fontsize=9, fontweight='bold', rotation=90
    )


def tracer_graphique(fichier, titre, dates, voies, instant_montant, instant_descente, config):
    couleurs = construire_couleurs(voies.keys())

    fig = plt.figure(figsize=tuple(config['figure_taille']))
    gs = fig.add_gridspec(2, 2, height_ratios=[1, 1.3], hspace=0.55, wspace=0.25)
    ax_montant = fig.add_subplot(gs[0, 0])
    ax_descente = fig.add_subplot(gs[0, 1])
    ax_bas = fig.add_subplot(gs[1, :])

    # Graphique du bas : ensemble des voies retenues, après filtrage/rognage
    for nom, valeurs in sorted(voies.items()):
        ax_bas.plot(dates, valeurs, color=couleurs[nom], label=nom, linewidth=1)
    if instant_montant is not None:
        ax_bas.axvline(instant_montant, color='red', linestyle='--', linewidth=1, alpha=0.6)
    if instant_descente is not None:
        ax_bas.axvline(instant_descente, color='red', linestyle='--', linewidth=1, alpha=0.6)
    ax_bas.set_xlabel("Temps")
    ax_bas.set_ylabel("Température (°C)")
    ax_bas.xaxis.set_major_formatter(mdates.DateFormatter('%Y/%m/%d %H:%M:%S'))
    for lbl in ax_bas.get_xticklabels():
        lbl.set_rotation(20)
        lbl.set_ha('right')
    ax_bas.grid(True, alpha=0.3)
    ax_bas.legend(loc='upper left', ncol=min(len(voies), 8), fontsize=7,
                   framealpha=0.9)

    # Zooms montée / descente
    _tracer_zoom(ax_montant, dates, voies, couleurs, instant_montant,
                 config['seuil_montant_T'], config['graphique_montant_dT'],
                 config['graphique_montant_dt'], "Montée")
    _tracer_zoom(ax_descente, dates, voies, couleurs, instant_descente,
                 config['seuil_descente_T'], config['graphique_descente_dT'],
                 config['graphique_descente_dt'], "Descente")

    fig.suptitle(titre or os.path.basename(fichier), fontsize=14, fontweight='bold', y=0.99)
    if instant_montant is not None and instant_descente is not None:
        duree = instant_descente - instant_montant
        fig.text(0.5, 0.95,
                  f"Durée montée → descente : {formater_duree(duree)}",
                  ha='center', fontsize=10, color='dimgray')

    fig.subplots_adjust(top=0.87, bottom=0.10, left=0.06, right=0.98)
    return fig


# =========================================================================
# Traitement d'un fichier
# =========================================================================

def analyser_fichier(fichier, config):
    print(f"📄 Lecture de {fichier}...")
    titre, dates, valeurs_brutes = lire_donnees_excel(fichier, config)
    print(f"   {len(dates)} échantillons, {len(valeurs_brutes)} voies détectées.")

    voies = filtrer_voies(valeurs_brutes, config['CH_ignore_dT'])
    voies_ignorees = sorted(set(valeurs_brutes) - set(voies))
    if voies_ignorees:
        print(f"   ⏭️  Voies ignorées (-OVER ou ΔT ≤ {config['CH_ignore_dT']}°) : "
              f"{', '.join(voies_ignorees)}")
    if not voies:
        print(f"   ⚠️ Aucune voie exploitable pour {fichier}, fichier ignoré.")
        return
    print(f"   ✅ Voies retenues : {', '.join(sorted(voies))}")

    dates_r, voies_r = rogner_debut_fin(dates, voies, config['time_ignore_dT'])
    print(f"   ✂️  Fenêtre temporelle retenue : {dates_r[0]} → {dates_r[-1]} "
          f"({len(dates_r)} échantillons)")

    instant_montant, _ = instant_seuil(
        dates_r, voies_r, config['seuil_montant_T'], 'montant',
        config['seuil_montant_first_or_last']
    )
    instant_descente, _ = instant_seuil(
        dates_r, voies_r, config['seuil_descente_T'], 'descente',
        config['seuil_descente_first_or_last']
    )

    if instant_montant is None:
        print(f"   ⚠️ Seuil de montée ({config['seuil_montant_T']}°) jamais franchi.")
    if instant_descente is None:
        print(f"   ⚠️ Seuil de descente ({config['seuil_descente_T']}°) jamais franchi.")
    if instant_montant is not None and instant_descente is not None:
        duree = instant_descente - instant_montant
        print(f"   ⏱️  Montée à {instant_montant} / Descente à {instant_descente} "
              f"/ Durée = {formater_duree(duree)}")

    fig = tracer_graphique(fichier, titre, dates_r, voies_r, instant_montant,
                            instant_descente, config)

    base, _ = os.path.splitext(fichier)
    fichier_sortie = f"{base}_{config['suffixe']}.{config['format_image']}"
    fig.savefig(fichier_sortie, dpi=config['dpi'])
    plt.close(fig)
    print(f"✅ Graphique généré -> {fichier_sortie}")


# =========================================================================
# Configuration (YAML) et arguments en ligne de commande
# =========================================================================

DEFAULT_CONFIG_FILE = "analyseTemperatureExcel.yaml"

DEFAULTS = {
    'CH_ignore_dT': 20.0,
    'time_ignore_dT': 5.0,
    'seuil_montant_T': 150.0,
    'seuil_montant_first_or_last': 'first',
    'seuil_descente_T': 150.0,
    'seuil_descente_first_or_last': 'last',
    'graphique_montant_dT': 30.0,
    'graphique_montant_dt': 300,
    'graphique_descente_dT': 30.0,
    'graphique_descente_dt': 300,
    'suffixe': 'Analyse',
    'format_image': 'png',
    'dpi': 150,
    'figure_taille': [14, 8],
    'feuille': None,
    'titre_graphique_label': 'Batch No.',
    'titre_graphique_col': 3,
}


def lire_parametres_yaml(fichier):
    if not os.path.exists(fichier):
        raise FileNotFoundError(f"Le fichier {fichier} est introuvable.")
    with open(fichier, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f) or {}


def construire_config(args):
    fichier_config = args.config or DEFAULT_CONFIG_FILE
    yaml_params = {}
    if os.path.exists(fichier_config):
        print(f"⚙️ Lecture des paramètres depuis {fichier_config}...")
        yaml_params = lire_parametres_yaml(fichier_config)
    elif args.config:
        # --config fourni explicitement mais introuvable -> erreur
        raise FileNotFoundError(f"Le fichier {fichier_config} est introuvable.")
    else:
        print(f"⚙️ Aucun fichier {fichier_config} trouvé, utilisation des valeurs par défaut.")

    def valeur(cli_val, cle_yaml):
        if cli_val is not None:
            return cli_val
        return yaml_params.get(cle_yaml, DEFAULTS[cle_yaml])

    config = {
        'CH_ignore_dT': float(valeur(args.ch_ignore_dt, 'CH_ignore_dT')),
        'time_ignore_dT': float(valeur(args.time_ignore_dt, 'time_ignore_dT')),
        'seuil_montant_T': float(valeur(args.seuil_montant, 'seuil_montant_T')),
        'seuil_montant_first_or_last': str(
            yaml_params.get('seuil_montant_first_or_last', DEFAULTS['seuil_montant_first_or_last'])
        ).strip().lower(),
        'seuil_descente_T': float(valeur(args.seuil_descente, 'seuil_descente_T')),
        'seuil_descente_first_or_last': str(
            yaml_params.get('seuil_descente_first_or_last', DEFAULTS['seuil_descente_first_or_last'])
        ).strip().lower(),
        'graphique_montant_dT': float(
            yaml_params.get('graphique_montant_dT', DEFAULTS['graphique_montant_dT'])
        ),
        'graphique_montant_dt': float(
            yaml_params.get('graphique_montant_dt', DEFAULTS['graphique_montant_dt'])
        ),
        'graphique_descente_dT': float(
            yaml_params.get('graphique_descente_dT', DEFAULTS['graphique_descente_dT'])
        ),
        'graphique_descente_dt': float(
            yaml_params.get('graphique_descente_dt', DEFAULTS['graphique_descente_dt'])
        ),
        'suffixe': str(yaml_params.get('suffixe', DEFAULTS['suffixe'])),
        'format_image': str(yaml_params.get('format_image', DEFAULTS['format_image'])),
        'dpi': int(yaml_params.get('dpi', DEFAULTS['dpi'])),
        'figure_taille': yaml_params.get('figure_taille', DEFAULTS['figure_taille']),
        'feuille': yaml_params.get('feuille', DEFAULTS['feuille']),
        'titre_graphique_label': str(
            yaml_params.get('titre_graphique_label', DEFAULTS['titre_graphique_label'])
        ),
        'titre_graphique_col': int(
            yaml_params.get('titre_graphique_col', DEFAULTS['titre_graphique_col'])
        ),
    }

    for cle in ('seuil_montant_first_or_last', 'seuil_descente_first_or_last'):
        if config[cle] not in ('first', 'last'):
            raise ValueError(f"{cle} doit valoir 'first' ou 'last' (valeur reçue : {config[cle]!r})")

    return config


def main():
    parser = argparse.ArgumentParser(
        description="Analyse des courbes de température d'un ou plusieurs fichiers Excel "
                    "et génération d'images de synthèse (montée / descente / vue globale)."
    )
    parser.add_argument("fichiers", nargs="*", help="Un ou plusieurs fichiers Excel (.xlsx) à traiter")
    parser.add_argument("--config", type=str, help="Chemin du fichier YAML de configuration")
    parser.add_argument("--ch-ignore-dt", dest="ch_ignore_dt", type=float,
                         help="Seuil d'amplitude (°C) en dessous duquel une voie est ignorée")
    parser.add_argument("--time-ignore-dt", dest="time_ignore_dt", type=float,
                         help="Seuil (°C) utilisé pour rogner le début/la fin des courbes")
    parser.add_argument("--seuil-montant", dest="seuil_montant", type=float,
                         help="Température de seuil pour l'instant de montée")
    parser.add_argument("--seuil-descente", dest="seuil_descente", type=float,
                         help="Température de seuil pour l'instant de descente")

    args = parser.parse_args()

    try:
        config = construire_config(args)
    except (FileNotFoundError, ValueError) as e:
        print(f"❌ Erreur de configuration : {e}")
        sys.exit(1)

    print("=== Paramètres d'analyse ===")
    for cle in ('CH_ignore_dT', 'time_ignore_dT', 'seuil_montant_T',
                'seuil_montant_first_or_last', 'seuil_descente_T',
                'seuil_descente_first_or_last', 'graphique_montant_dT',
                'graphique_montant_dt', 'graphique_descente_dT',
                'graphique_descente_dt', 'suffixe'):
        print(f"{cle} = {config[cle]}")
    print()

    if args.fichiers:
        print(f"🔍 {len(args.fichiers)} fichier(s) Excel fourni(s) :")
        for fichier in args.fichiers:
            try:
                analyser_fichier(fichier, config)
            except Exception as e:
                print(f"❌ Erreur lors du traitement de {fichier} : {e}")
    else:
        print("Traitement des fichiers .xlsx (hors fichiers déjà suffixés) du répertoire actuel.")
        suffixe_bas = f"_{config['suffixe']}".lower()
        fichiers_xlsx = [
            f for f in os.listdir('.')
            if f.lower().endswith('.xlsx') and not f.lower()[:-5].endswith(suffixe_bas)
        ]
        if not fichiers_xlsx:
            print("⚠️ Aucun fichier .xlsx trouvé dans le répertoire actuel.")
            return

        print(f"🔍 {len(fichiers_xlsx)} fichier(s) .xlsx trouvé(s) :")
        for f in fichiers_xlsx:
            print(f" → {f}")
        print("\n⏳ Traitement en cours...\n")

        for fichier in fichiers_xlsx:
            try:
                analyser_fichier(fichier, config)
            except Exception as e:
                print(f"❌ Erreur lors du traitement de {fichier} : {e}")

        print("\n✅ Tous les fichiers .xlsx ont été traités.")


if __name__ == "__main__":
    main()
